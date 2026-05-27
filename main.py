from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from pydantic import BaseModel
from datetime import date, time, datetime, timedelta
from typing import Optional
import asyncpg
import os
import resend
from dotenv import load_dotenv
from fastapi.responses import StreamingResponse
import openpyxl
from io import BytesIO
import bcrypt
import httpx
from main_qfa import qfa_app, init_qfa_pool

load_dotenv()

app = FastAPI(title="ReservaTuEspacio")

async def obtener_tipo_cambio(db) -> float:
    """Lee el tipo de cambio USD/ARS desde la tabla configuracion_global."""
    try:
        row = await db.fetchrow(
            "SELECT valor FROM configuracion_global WHERE clave = 'valor_dolar_oficial'"
        )
        if row:
            return float(row["valor"])
    except Exception as e:
        print(f"Error al leer tipo de cambio: {e}")
    return float(os.getenv("TC_FALLBACK_USD_ARS", "1415"))

resend.api_key = os.getenv("RESEND_API_KEY")

# Rate limiting
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Orígenes permitidos — agregar cualquier dominio nuevo que necesite acceso
ALLOWED_ORIGINS = [
    "https://reservatuespacio.com",
    "https://www.reservatuespacio.com",
    "https://claudiaacreativity.github.io",
    "https://gestionateia.com",
    "https://www.gestionateia.com",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "X-Tenant-Slug", "X-Admin-Token", "X-Superadmin-Token", "Authorization"],
)

app.mount("/quefiestaapp", qfa_app)

pool = None

@app.on_event("startup")
async def startup():
    global pool
    pool = await asyncpg.create_pool(
        host=os.getenv("DB_HOST"),
        port=int(os.getenv("DB_PORT")),
        database=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        min_size=1,
        max_size=5
    )
    await init_qfa_pool()

@app.on_event("shutdown")
async def shutdown():
    await pool.close()

async def get_db():
    return await pool.acquire()

async def release_db(conn):
    await pool.release(conn)

# ===== MULTI-TENANT =====
async def get_tenant(request: Request, db):
    host = request.headers.get("host", "")
    # Extraer subdominio: prueba.reservas.com -> prueba
    # En desarrollo local usamos header X-Tenant-Slug
    slug = request.headers.get("X-Tenant-Slug", "")
    if not slug:
        parts = host.split(".")
        if len(parts) >= 3:
            slug = parts[0]
        else:
            # Fallback para desarrollo: usar tenant de prueba
            slug = "prueba"

    tenant = await db.fetchrow(
        "SELECT * FROM tenants WHERE slug = $1 AND activo = TRUE",
        slug
    )
    if not tenant:
        raise HTTPException(status_code=404, detail="Organización no encontrada")

    # Verificar si el trial venció y no tiene suscripción activa
    if not tenant["suscripcion_activa"] and tenant["trial_hasta"] < date.today():
        raise HTTPException(status_code=402, detail="El período de prueba ha vencido. Por favor suscribite para continuar.")

    return dict(tenant)

# ===== EMAIL =====
def enviar_email(destinatario: str, asunto: str, cuerpo: str):
    try:
        resend.Emails.send({
            "from": "ReservaTuEspacio <hola@reservatuespacio.com>",
            "to": destinatario,
            "subject": asunto,
            "html": cuerpo
        })
    except Exception as e:
        print(f"Error al enviar email: {e}")


async def enviar_email_aviso_vencimiento(email: str, nombre_tenant: str, dias: int, fecha_vence: date):
    """Avisa al tenant que su suscripción vence en X días."""
    fecha_str = fecha_vence.strftime("%d/%m/%Y")
    asunto = f"Tu suscripción vence en {dias} día{'s' if dias > 1 else ''} — ReservaTuEspacio"
    cuerpo = f"""
    <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
        <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
            <h2 style="color:#71D997; margin:0">Recordatorio de renovación</h2>
        </div>
        <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
            <p>Hola <b>{nombre_tenant}</b>,</p>
            <p>Te recordamos que tu suscripción a <b>ReservaTuEspacio</b> vence el
               <b>{fecha_str}</b> (en {dias} día{'s' if dias > 1 else ''}).</p>
            <p>Para renovar, ingresá a la landing y realizá el pago de tu plan.</p>
            <div style="margin:24px 0; text-align:center">
                <a href="https://reservatuespacio.com/landing.html"
                   style="background:#71D997; color:#2C3E50; padding:12px 28px;
                          border-radius:50px; text-decoration:none; font-weight:bold">
                    Renovar ahora →
                </a>
            </div>
            <p style="color:#888; font-size:13px">
                Si ya realizaste el pago, podés ignorar este mensaje.<br>
                ¿Tenés dudas?
                <a href="https://reservatuespacio.com/soporte.html" style="color:#888">Centro de soporte →</a>
            </p>
        </div>
    </div>
    """
    try:
        resend.Emails.send({
            "from": "ReservaTuEspacio <hola@reservatuespacio.com>",
            "to": email,
            "subject": asunto,
            "html": cuerpo
        })
    except Exception as e:
        print(f"[EMAIL AVISO VENCIMIENTO] Error enviando a {email}: {e}")


async def enviar_email_cuenta_bloqueada(email: str, nombre_tenant: str, fecha_vence: date):
    """Notifica al tenant que su cuenta fue bloqueada por vencimiento."""
    fecha_str = fecha_vence.strftime("%d/%m/%Y")
    asunto = "Tu suscripción venció — cuenta suspendida temporalmente"
    cuerpo = f"""
    <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
        <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
            <h2 style="color:#E74C3C; margin:0">Cuenta suspendida</h2>
        </div>
        <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
            <p>Hola <b>{nombre_tenant}</b>,</p>
            <p>Tu suscripción a <b>ReservaTuEspacio</b> venció el <b>{fecha_str}</b>
               y tu cuenta fue suspendida temporalmente.</p>
            <p><b>Tus datos están seguros.</b> Para reactivar tu cuenta, realizá el pago de renovación:</p>
            <div style="margin:24px 0; text-align:center">
                <a href="https://reservatuespacio.com/landing.html"
                   style="background:#71D997; color:#2C3E50; padding:12px 28px;
                          border-radius:50px; text-decoration:none; font-weight:bold">
                    Reactivar cuenta →
                </a>
            </div>
            <p style="color:#888; font-size:13px">
                ¿Necesitás ayuda?
                <a href="https://reservatuespacio.com/soporte.html" style="color:#888">Centro de soporte →</a>
            </p>
        </div>
    </div>
    """
    try:
        resend.Emails.send({
            "from": "ReservaTuEspacio <hola@reservatuespacio.com>",
            "to": email,
            "subject": asunto,
            "html": cuerpo
        })
    except Exception as e:
        print(f"[EMAIL CUENTA BLOQUEADA] Error enviando a {email}: {e}")

import secrets
import string

def generar_password_temporal(largo=12):
    """Genera una contraseña aleatoria segura."""
    chars = string.ascii_letters + string.digits + "!@#$%"
    while True:
        pwd = "".join(secrets.choice(chars) for _ in range(largo))
        if (any(c.isupper() for c in pwd) and
            any(c.islower() for c in pwd) and
            any(c.isdigit() for c in pwd)):
            return pwd


def email_bienvenida_tenant(nombre_admin, nombre_tenant, slug, email, password_temp, trial_hasta):
    """Genera el HTML del email de bienvenida con credenciales de acceso."""
    return f"""
    <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
        <div style="background:#2C3E50; padding:28px 32px; border-radius:12px 12px 0 0; text-align:center">
            <h1 style="color:#71D997; margin:0; font-size:26px">\u00a1Bienvenido a ReservaTuEspacio!</h1>
        </div>
        <div style="background:#F9F9FB; padding:28px 32px; border-radius:0 0 12px 12px">
            <p style="font-size:15px; color:#2C3E50">Hola <b>{nombre_admin}</b>,</p>
            <p style="color:#4A5568; line-height:1.7">
                Tu cuenta para <b>{nombre_tenant}</b> fue creada exitosamente.
                Ten\u00e9s <b>30 d\u00edas de prueba gratuita</b> hasta el <b>{trial_hasta}</b>.
            </p>
            <div style="background:white; border:2px solid #71D997; border-radius:10px; padding:20px; margin:24px 0">
                <p style="margin:0 0 12px; font-weight:bold; color:#2C3E50; font-size:15px">\U0001f511 Tus datos de acceso</p>
                <table style="width:100%; border-collapse:collapse">
                    <tr>
                        <td style="padding:8px 0; color:#7f8c8d; font-size:13px; width:120px">Panel admin:</td>
                        <td style="padding:8px 0">
                            <a href="https://reservatuespacio.com/admin/{slug}"
                               style="color:#71D997; font-weight:bold; font-size:14px">
                                reservatuespacio.com/admin/{slug}
                            </a>
                        </td>
                    </tr>
                    <tr style="background:#f9f9fb">
                        <td style="padding:8px 0; color:#7f8c8d; font-size:13px">Email:</td>
                        <td style="padding:8px 0; color:#2C3E50; font-weight:bold; font-size:14px">{email}</td>
                    </tr>
                    <tr>
                        <td style="padding:8px 0; color:#7f8c8d; font-size:13px">Contrase\u00f1a:</td>
                        <td style="padding:8px 0; color:#2C3E50; font-weight:bold; font-size:14px; font-family:monospace">{password_temp}</td>
                    </tr>
                    <tr style="background:#f9f9fb">
                        <td style="padding:8px 0; color:#7f8c8d; font-size:13px">Tu URL p\u00fablica:</td>
                        <td style="padding:8px 0">
                            <a href="https://reservatuespacio.com/{slug}"
                               style="color:#71D997; font-weight:bold; font-size:14px">
                                reservatuespacio.com/{slug}
                            </a>
                        </td>
                    </tr>
                </table>
            </div>
            <p style="color:#E74C3C; font-size:13px; font-weight:bold">
                \u26a0\ufe0f Por seguridad, te recomendamos cambiar tu contrase\u00f1a despu\u00e9s del primer ingreso
                desde el panel \u2192 Mi perfil.
            </p>
            <div style="text-align:center; margin-top:24px">
                <a href="https://reservatuespacio.com/admin/{slug}"
                   style="background:#71D997; color:#2C3E50; padding:14px 32px;
                          border-radius:50px; text-decoration:none; font-weight:bold; font-size:15px">
                    Ir al panel de administraci\u00f3n \u2192
                </a>
            </div>
            <p style="text-align:center; margin-top:16px; font-size:13px; color:#888">
                <a href="https://reservatuespacio.com/manual_reservatuespacio.pdf" style="color:#2C3E50; font-weight:bold">Manual de uso →</a>
                &nbsp;&nbsp;|&nbsp;&nbsp;
                ¿Tenés dudas?
                <a href="https://reservatuespacio.com/soporte.html" style="color:#888">Centro de soporte →</a>
            </p>
        </div>
    </div>
    """



# Modelos
class ReservaCreate(BaseModel):
    espacio_id: str
    usuario_id: Optional[str] = None        # None si es invitado
    fecha: date
    hora_inicio: time
    hora_fin: time
    # Datos de invitado (cuando usuario_id es None)
    invitado_nombre: Optional[str] = None
    invitado_email: Optional[str] = None
    invitado_whatsapp: Optional[str] = None
    # Recursos solicitados
    recursos: Optional[list] = None         # [{"recurso_id": "...", "cantidad": 2}]
    # Comprobante de transferencia
    comprobante_url: Optional[str] = None

class CancelarReserva(BaseModel):
    reserva_id: str
    usuario_id: str

class FechaBloqueada(BaseModel):
    fecha: date
    motivo: str

class EspacioCreate(BaseModel):
    nombre: str
    capacidad: Optional[int] = None
    edificio_id: int

class HorarioUpdate(BaseModel):
    habilitado: bool
    hora_apertura: Optional[time] = None
    hora_cierre: Optional[time] = None


# ===== ADMIN LOGIN =====

class AdminLogin(BaseModel):
    password: str

@app.post("/admin/login")
@limiter.limit("5/minute")
async def admin_login(request: Request, datos: AdminLogin):
    import secrets
    password_correcta = os.getenv("ADMIN_PASSWORD", "sL2#di!KBw")
    if datos.password != password_correcta:
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    # Generar token simple
    token = secrets.token_hex(32)
    return {"token": token}

@app.post("/admin/verificar")
async def admin_verificar(request: Request):
    token = request.headers.get("X-Admin-Token", "")
    if not token or len(token) != 64:
        raise HTTPException(status_code=401, detail="Token inválido")
    return {"valido": True}

# ===== ADMIN DEL TENANT =====

class AdminTenantLogin(BaseModel):
    email: str
    password: str

class SetPassword(BaseModel):
    password: str

@app.post("/admin/tenant/login")
@limiter.limit("5/minute")
async def admin_tenant_login(request: Request, datos: AdminTenantLogin):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        usuario = await db.fetchrow(
            "SELECT * FROM usuarios WHERE email = $1 AND tenant_id = $2 AND rol = 'admin' AND activo = TRUE",
            datos.email, tenant["id"]
        )
        if not usuario:
            raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")
        if not usuario["password_hash"]:
            raise HTTPException(status_code=401, detail="Este usuario no tiene contraseña configurada")
        password_ok = bcrypt.checkpw(datos.password.encode(), usuario["password_hash"].encode())
        if not password_ok:
            raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")
        import secrets
        token = secrets.token_hex(32)
        return {
            "token": token,
            "usuario_id": str(usuario["id"]),
            "nombre": usuario["nombre"],
            "email": usuario["email"]
        }
    finally:
        await release_db(db)

@app.post("/usuarios/{usuario_id}/set-password")
async def set_admin_password(request: Request, usuario_id: str, datos: SetPassword):
    """Asigna contraseña a un usuario y lo promueve a admin."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        usuario = await db.fetchrow(
            "SELECT * FROM usuarios WHERE id = $1 AND tenant_id = $2",
            usuario_id, tenant["id"]
        )
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        if len(datos.password) < 8:
            raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")
        password_hash = bcrypt.hashpw(datos.password.encode(), bcrypt.gensalt()).decode()
        await db.execute(
            "UPDATE usuarios SET password_hash = $1, rol = 'admin' WHERE id = $2 AND tenant_id = $3",
            password_hash, usuario_id, tenant["id"]
        )
        return {"mensaje": f"{usuario['nombre']} ahora es administrador"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


class AdminPerfilUpdate(BaseModel):
    nombre: Optional[str] = None
    email: Optional[str] = None
    password_actual: Optional[str] = None
    password_nueva: Optional[str] = None

@app.patch("/admin/perfil")
async def actualizar_perfil_admin(request: Request, datos: AdminPerfilUpdate):
    """Permite al admin del tenant actualizar su propio nombre, email y contraseña."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        # Identificar al admin por el header X-Admin-Email
        email_actual = request.headers.get("X-Admin-Email", "")
        if not email_actual:
            raise HTTPException(status_code=400, detail="Se requiere el header X-Admin-Email")
        usuario = await db.fetchrow(
            "SELECT * FROM usuarios WHERE email = $1 AND tenant_id = $2 AND rol = 'admin'",
            email_actual, tenant["id"]
        )
        if not usuario:
            raise HTTPException(status_code=404, detail="Administrador no encontrado")

        # Actualizar nombre
        if datos.nombre:
            await db.execute(
                "UPDATE usuarios SET nombre = $1 WHERE id = $2",
                datos.nombre, usuario["id"]
            )

        # Actualizar email
        if datos.email and datos.email != email_actual:
            existente = await db.fetchrow(
                "SELECT id FROM usuarios WHERE email = $1 AND tenant_id = $2",
                datos.email, tenant["id"]
            )
            if existente:
                raise HTTPException(status_code=400, detail="Ese email ya está en uso")
            await db.execute(
                "UPDATE usuarios SET email = $1 WHERE id = $2",
                datos.email, usuario["id"]
            )

        # Actualizar contraseña
        if datos.password_nueva:
            if not datos.password_actual:
                raise HTTPException(status_code=400, detail="Debés ingresar tu contraseña actual")
            if not bcrypt.checkpw(datos.password_actual.encode(), usuario["password_hash"].encode()):
                raise HTTPException(status_code=401, detail="La contraseña actual es incorrecta")
            if len(datos.password_nueva) < 8:
                raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 8 caracteres")
            nuevo_hash = bcrypt.hashpw(datos.password_nueva.encode(), bcrypt.gensalt()).decode()
            await db.execute(
                "UPDATE usuarios SET password_hash = $1 WHERE id = $2",
                nuevo_hash, usuario["id"]
            )

        return {"mensaje": "Perfil actualizado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


import json as json_module

class TenantPerfilUpdate(BaseModel):
    logo_url: Optional[str] = None
    favicon_url: Optional[str] = None
    foto_url: Optional[str] = None
    color_primario: Optional[str] = None
    color_secundario: Optional[str] = None
    direccion: Optional[str] = None
    whatsapp: Optional[str] = None
    redes_sociales: Optional[list] = None
    foto_posicion: Optional[str] = None
    modalidad_cobro: Optional[str] = None
    politica_cancelacion: Optional[str] = None
    horas_cancelacion: Optional[int] = None
    registro_requerido: Optional[bool] = None
    transferencia_habilitada: Optional[bool] = None
    transferencia_alias: Optional[str] = None
    transferencia_mensaje: Optional[str] = None

@app.patch("/tenant/perfil")
async def actualizar_perfil_tenant(request: Request, datos: TenantPerfilUpdate):
    """Permite al admin actualizar los datos de personalización del tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        campos = []
        valores = []
        i = 1
        if datos.logo_url is not None:
            campos.append(f"logo_url = ${i}"); valores.append(datos.logo_url); i += 1
        if datos.favicon_url is not None:
            campos.append(f"favicon_url = ${i}"); valores.append(datos.favicon_url); i += 1
        if datos.foto_url is not None:
            campos.append(f"foto_url = ${i}"); valores.append(datos.foto_url); i += 1
        if datos.color_primario is not None:
            campos.append(f"color_primario = ${i}"); valores.append(datos.color_primario); i += 1
        if datos.color_secundario is not None:
            campos.append(f"color_secundario = ${i}"); valores.append(datos.color_secundario); i += 1
        if datos.direccion is not None:
            campos.append(f"direccion = ${i}"); valores.append(datos.direccion); i += 1
        if datos.whatsapp is not None:
            campos.append(f"whatsapp = ${i}"); valores.append(datos.whatsapp); i += 1
        if datos.redes_sociales is not None:
            campos.append(f"redes_sociales = ${i}"); valores.append(json_module.dumps(datos.redes_sociales)); i += 1
        if datos.foto_posicion is not None:
            campos.append(f"foto_posicion = ${i}"); valores.append(datos.foto_posicion); i += 1
        if datos.modalidad_cobro is not None:
            campos.append(f"modalidad_cobro = ${i}"); valores.append(datos.modalidad_cobro); i += 1
        if datos.politica_cancelacion is not None:
            campos.append(f"politica_cancelacion = ${i}"); valores.append(datos.politica_cancelacion); i += 1
        if datos.horas_cancelacion is not None:
            campos.append(f"horas_cancelacion = ${i}"); valores.append(datos.horas_cancelacion); i += 1
        if datos.registro_requerido is not None:
            campos.append(f"registro_requerido = ${i}"); valores.append(datos.registro_requerido); i += 1
        if datos.transferencia_habilitada is not None:
            campos.append(f"transferencia_habilitada = ${i}"); valores.append(datos.transferencia_habilitada); i += 1
        if datos.transferencia_alias is not None:
            campos.append(f"transferencia_alias = ${i}"); valores.append(datos.transferencia_alias); i += 1
        if datos.transferencia_mensaje is not None:
            campos.append(f"transferencia_mensaje = ${i}"); valores.append(datos.transferencia_mensaje); i += 1
        if not campos:
            return {"mensaje": "No hay cambios para guardar"}
        valores.append(tenant["id"])
        await db.execute(
            f"UPDATE tenants SET {', '.join(campos)} WHERE id = ${i}",
            *valores
        )
        return {"mensaje": "Perfil del tenant actualizado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


@app.post("/tenant/imagen")
async def subir_imagen_tenant(request: Request):
    """Recibe una imagen en base64 y la guarda en Supabase Storage."""
    import base64
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        body = await request.json()
        imagen_b64 = body.get("imagen")
        tipo = body.get("tipo", "logo")  # logo, favicon, foto
        mime = body.get("mime", "image/png")
        
        if not imagen_b64:
            raise HTTPException(status_code=400, detail="Se requiere el campo imagen en base64")
        
        import httpx
        imagen_bytes = base64.b64decode(imagen_b64)
        extension = mime.split("/")[-1].replace("jpeg", "jpg")
        filename = f"{tenant['slug']}/{tipo}.{extension}"
        
        supabase_url = f"https://okkwfaouqdnbityotnje.supabase.co/storage/v1/object/logos/{filename}"
        supabase_key = os.getenv("SUPABASE_SERVICE_KEY", "")
        
        if not supabase_key:
            raise HTTPException(status_code=500, detail="SUPABASE_SERVICE_KEY no configurada")
        
        async with httpx.AsyncClient() as client:
            res = await client.put(
                supabase_url,
                content=imagen_bytes,
                headers={
                    "Authorization": f"Bearer {supabase_key}",
                    "Content-Type": mime,
                    "x-upsert": "true"
                }
            )
        
        if res.status_code not in (200, 201):
            raise HTTPException(status_code=500, detail=f"Error al subir imagen: {res.text}")
        
        url_publica = f"https://okkwfaouqdnbityotnje.supabase.co/storage/v1/object/public/logos/{filename}"
        return {"url": url_publica}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

# ===== ENDPOINTS =====

@app.api_route("/", methods=["GET", "HEAD"])
async def inicio():
    return {"mensaje": "ReservaTuEspacio funcionando"}

@app.get("/tenant")
async def info_tenant(request: Request):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        return {
            "nombre": tenant["nombre"],
            "slug": tenant["slug"],
            "logo_url": tenant["logo_url"],
            "favicon_url": tenant["favicon_url"],
            "color_primario": tenant["color_primario"],
            "color_secundario": tenant["color_secundario"],
            "plan": tenant["plan_id"],
            "direccion": tenant["direccion"],
            "whatsapp": tenant["whatsapp"],
            "redes_sociales": (json_module.loads(tenant["redes_sociales"]) if isinstance(tenant["redes_sociales"], str) else tenant["redes_sociales"]) or [],
            "foto_url": tenant["foto_url"],
            "foto_posicion": tenant["foto_posicion"],
            "modalidad_cobro": tenant["modalidad_cobro"],
            "politica_cancelacion": tenant["politica_cancelacion"],
            "horas_cancelacion": tenant["horas_cancelacion"],
            "registro_requerido": bool(tenant.get("registro_requerido", False)),
            "transferencia_habilitada": bool(tenant.get("transferencia_habilitada", False)),
            "transferencia_alias": tenant.get("transferencia_alias") or "",
            "transferencia_mensaje": tenant.get("transferencia_mensaje") or "",
        }
    finally:
        await release_db(db)

@app.get("/espacios")
async def listar_espacios(request: Request):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        espacios = await db.fetch("""
            SELECT a.*, COALESCE(e.nombre, a.edificio, '') as nombre_edificio
            FROM aulas a
            LEFT JOIN edificios e ON a.edificio_id = e.id
            WHERE a.activa = TRUE AND a.tenant_id = $1
        """, tenant["id"])
        return [dict(e) for e in espacios]
    finally:
        await release_db(db)

@app.post("/espacios")
async def crear_espacio(request: Request, espacio: EspacioCreate):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        result = await db.fetchrow(
            """INSERT INTO aulas (nombre, capacidad, edificio_id, tenant_id)
               VALUES ($1, $2, $3, $4) RETURNING id""",
            espacio.nombre, espacio.capacidad, espacio.edificio_id, tenant["id"]
        )
        return {"mensaje": "Espacio creado", "id": str(result["id"])}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.patch("/espacios/{espacio_id}")
async def toggle_espacio(request: Request, espacio_id: str, datos: dict):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        await db.execute(
            "UPDATE aulas SET activa = $1 WHERE id = $2 AND tenant_id = $3",
            datos["activa"], espacio_id, tenant["id"]
        )
        return {"mensaje": "Espacio actualizado"}
    finally:
        await release_db(db)

@app.get("/disponibilidad/{espacio_id}/{fecha}")
async def consultar_disponibilidad(request: Request, espacio_id: str, fecha: date):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reservas = await db.fetch(
            """SELECT hora_inicio, hora_fin, estado FROM reservas
               WHERE aula_id=$1 AND fecha=$2
                 AND estado IN ('activa','pending_payment')
                 AND tenant_id=$3""",
            espacio_id, fecha, tenant["id"]
        )
        return [dict(r) for r in reservas]
    finally:
        await release_db(db)

@app.post("/reservas")
async def crear_reserva(request: Request, reserva: ReservaCreate):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        tid = tenant["id"]

        # Validar registro requerido
        if tenant.get("registro_requerido"):
            if not reserva.usuario_id:
                raise HTTPException(
                    status_code=403,
                    detail="Este espacio requiere que estés registrado para hacer una reserva. Contactá al administrador."
                )
            usuario = await db.fetchrow(
                "SELECT id FROM usuarios WHERE id = $1 AND tenant_id = $2 AND activo = TRUE",
                reserva.usuario_id, tid
            )
            if not usuario:
                raise HTTPException(
                    status_code=403,
                    detail="Tu usuario no está registrado o no está activo. Contactá al administrador."
                )

        ahora = datetime.now()
        fecha_hora_inicio = datetime.combine(reserva.fecha, reserva.hora_inicio)
        if fecha_hora_inicio < ahora:
            raise HTTPException(
                status_code=400,
                detail="La fecha y hora de reserva debe ser posterior a la actual"
            )

        fecha_bloqueada = await db.fetchrow(
            "SELECT motivo FROM fechas_bloqueadas WHERE fecha = $1 AND tenant_id = $2",
            reserva.fecha, tid
        )
        if fecha_bloqueada:
            raise HTTPException(
                status_code=400,
                detail=f"No se puede reservar en esa fecha: {fecha_bloqueada['motivo']}"
            )

        dia_semana = reserva.fecha.weekday()
        config = await db.fetchrow(
            "SELECT * FROM configuracion_horarios WHERE dia_semana = $1 AND tenant_id = $2",
            dia_semana, tid
        )

        if not config or not config["habilitado"]:
            raise HTTPException(
                status_code=400,
                detail=f"No hay disponibilidad los {config['nombre_dia']}s"
            )

        if reserva.hora_inicio < config["hora_apertura"]:
            raise HTTPException(
                status_code=400,
                detail=f"El horario de apertura es a las {config['hora_apertura'].strftime('%H:%M')}"
            )

        if reserva.hora_fin > config["hora_cierre"]:
            raise HTTPException(
                status_code=400,
                detail=f"El horario de cierre es a las {config['hora_cierre'].strftime('%H:%M')}"
            )

        # Validar: usuario registrado o invitado con datos completos
        if not reserva.usuario_id:
            if not reserva.invitado_nombre or not reserva.invitado_email or not reserva.invitado_whatsapp:
                raise HTTPException(status_code=400, detail="Para reservar sin cuenta debés ingresar nombre, email y WhatsApp")

        # Validar disponibilidad de recursos si se solicitaron
        recursos_a_reservar = []
        if reserva.recursos:
            for item in reserva.recursos:
                recurso_id = item.get("recurso_id")
                cantidad_solicitada = int(item.get("cantidad", 1))
                recurso = await db.fetchrow(
                    "SELECT id, nombre, cantidad_total FROM recursos WHERE id = $1 AND tenant_id = $2 AND activo = TRUE",
                    recurso_id, tid
                )
                if not recurso:
                    raise HTTPException(status_code=404, detail=f"Recurso no encontrado")
                # Calcular cantidad ya reservada en el mismo horario
                en_uso = await db.fetchval(
                    """SELECT COALESCE(SUM(rr.cantidad), 0)
                       FROM recursos_reservas rr
                       JOIN reservas r ON rr.reserva_id = r.id
                       WHERE rr.recurso_id = $1
                       AND r.fecha = $2
                       AND r.estado = 'activa'
                       AND r.tenant_id = $3
                       AND r.hora_inicio < $4 AND r.hora_fin > $5""",
                    recurso_id, reserva.fecha, tid, reserva.hora_fin, reserva.hora_inicio
                )
                disponibles = recurso["cantidad_total"] - int(en_uso)
                if disponibles <= 0:
                    raise HTTPException(
                        status_code=400,
                        detail=f"No hay {recurso['nombre']} disponibles en ese horario"
                    )
                cantidad_final = min(cantidad_solicitada, disponibles)
                recursos_a_reservar.append({
                    "recurso_id": recurso_id,
                    "cantidad": cantidad_final,
                    "solicitada": cantidad_solicitada,
                    "disponibles": disponibles,
                    "nombre": recurso["nombre"]
                })

        # Validar superposición (incluye pending_payment para evitar doble reserva)
        solapamiento = await db.fetchrow(
            """SELECT id FROM reservas
               WHERE aula_id = $1 AND fecha = $2
                 AND estado IN ('activa', 'pending_payment')
                 AND tenant_id = $3
                 AND hora_inicio < $4 AND hora_fin > $5""",
            reserva.espacio_id, reserva.fecha, tid,
            reserva.hora_fin, reserva.hora_inicio
        )
        if solapamiento:
            raise HTTPException(
                status_code=400,
                detail="Ese horario ya está reservado o está siendo procesado"
            )

        result = await db.fetchrow(
            """INSERT INTO reservas (aula_id, usuario_id, fecha, hora_inicio, hora_fin, tenant_id,
                                     invitado_nombre, invitado_email, invitado_whatsapp, comprobante_url)
               VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10) RETURNING id""",
            reserva.espacio_id, reserva.usuario_id, reserva.fecha,
            reserva.hora_inicio, reserva.hora_fin, tid,
            reserva.invitado_nombre, reserva.invitado_email, reserva.invitado_whatsapp,
            reserva.comprobante_url
        )
        reserva_id = result["id"]

        # Guardar recursos reservados
        for item in recursos_a_reservar:
            await db.execute(
                "INSERT INTO recursos_reservas (reserva_id, recurso_id, cantidad) VALUES ($1, $2, $3)",
                reserva_id, item["recurso_id"], item["cantidad"]
            )

        usuario = await db.fetchrow(
            "SELECT email, nombre FROM usuarios WHERE id = $1 AND tenant_id = $2",
            reserva.usuario_id, tid
        ) if reserva.usuario_id else None
        email_destino = usuario["email"] if usuario else reserva.invitado_email
        nombre_destino = usuario["nombre"] if usuario else reserva.invitado_nombre
        espacio = await db.fetchrow(
            "SELECT nombre FROM aulas WHERE id = $1 AND tenant_id = $2",
            reserva.espacio_id, tid
        )
        # Armar recursos HTML para el email
        recursos_html = ""
        if recursos_a_reservar:
            items_html = "".join([
                f'<tr><td style="padding:10px 14px; color:#4A5568">{r["nombre"]}</td><td style="padding:10px 14px; color:#4A5568">{r["cantidad"]}</td></tr>'
                for r in recursos_a_reservar
            ])
            recursos_html = f"""
            <div style="margin-top:16px">
                <p style="font-weight:bold; color:#2C3E50; margin-bottom:8px">Recursos reservados:</p>
                <table style="border-collapse:collapse; width:100%; background:white; border-radius:8px; overflow:hidden">
                    <tr style="background:#f0f4f8">
                        <td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Recurso</td>
                        <td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Cantidad</td>
                    </tr>
                    {items_html}
                </table>
            </div>"""

        if email_destino:
            # Build politica section if exists
            politica_html = ""
            if tenant.get("politica_cancelacion"):
                modalidades = {
                    "por_reserva": "Pago por reserva",
                    "mes_adelantado": "Pago mensual adelantado",
                    "mes_vencido": "Facturación a mes vencido",
                    "sin_cobro": "Sin cobro"
                }
                modalidad_texto = modalidades.get(tenant.get("modalidad_cobro", "sin_cobro"), "")
                politica_html = f"""
                <div style="margin-top:20px; background:#f9f9fb; border-left:4px solid #71D997; padding:14px 16px; border-radius:0 8px 8px 0">
                    <p style="font-weight:bold; color:#2C3E50; margin:0 0 6px">Política de cancelación y reembolso</p>
                    {f'<p style="font-size:13px; color:#4A5568; margin:0 0 6px"><b>Modalidad de cobro:</b> {modalidad_texto}</p>' if modalidad_texto and modalidad_texto != "Sin cobro" else ""}
                    <p style="font-size:13px; color:#4A5568; margin:0; line-height:1.6">{tenant["politica_cancelacion"]}</p>
                </div>"""

            enviar_email(
                email_destino,
                "✅ Reserva confirmada — " + tenant["nombre"],
                f"""
                <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
                    <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
                        <h2 style="color:#71D997; margin:0">¡Reserva confirmada!</h2>
                    </div>
                    <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
                        <p style="color:#2C3E50; margin-bottom:16px">Hola <b>{nombre_destino}</b>, tu reserva fue registrada correctamente.</p>
                        <table style="border-collapse:collapse; width:100%; background:white; border-radius:8px; overflow:hidden">
                            <tr style="background:#f0f4f8"><td style="padding:10px 14px; font-weight:bold; color:#2C3E50; width:120px">Espacio</td><td style="padding:10px 14px; color:#4A5568">{espacio['nombre']}</td></tr>
                            <tr><td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Fecha</td><td style="padding:10px 14px; color:#4A5568">{reserva.fecha.strftime('%d/%m/%Y')}</td></tr>
                            <tr style="background:#f0f4f8"><td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Horario</td><td style="padding:10px 14px; color:#4A5568">{reserva.hora_inicio.strftime('%H:%M')} - {reserva.hora_fin.strftime('%H:%M')}</td></tr>
                            <tr><td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Organización</td><td style="padding:10px 14px; color:#4A5568">{tenant['nombre']}</td></tr>
                        </table>
                        {recursos_html}
                        {politica_html}
                        <p style="margin-top:20px; font-size:13px; color:#888; text-align:center">
                            <a href="https://reservatuespacio.com/faq-usuarios.html" style="color:#2C3E50; font-weight:bold">Preguntas frecuentes →</a>
                            &nbsp;&nbsp;|&nbsp;&nbsp;
                            <a href="https://reservatuespacio.com/soporte.html" style="color:#888">¿Necesitás ayuda? Centro de soporte →</a>
                        </p>
                    </div>
                </div>
                """
            )
        # Advertencia si no se pudo satisfacer toda la cantidad de recursos
        avisos = [
            f"Solo quedaban {r['disponibles']} {r['nombre']} disponibles (pediste {r['solicitada']})"
            for r in recursos_a_reservar if r["cantidad"] < r["solicitada"]
        ]
        return {"mensaje": "Reserva creada", "id": str(reserva_id), "avisos": avisos}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.delete("/reservas/{reserva_id}")
async def cancelar_reserva(request: Request, reserva_id: str, datos: CancelarReserva):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reserva = await db.fetchrow(
            """SELECT id, usuario_id, invitado_email, invitado_nombre,
                      mp_payment_id, monto, fecha, hora_inicio, hora_fin
               FROM reservas
               WHERE id=$1 AND estado='activa' AND tenant_id=$2""",
            reserva_id, tenant["id"]
        )
        if not reserva:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")

        # Validar identidad: usuario registrado o invitado
        if reserva["usuario_id"]:
            if str(reserva["usuario_id"]) != datos.usuario_id:
                raise HTTPException(status_code=403, detail="Solo el usuario que creó la reserva puede cancelarla")
        elif reserva["invitado_email"]:
            if reserva["invitado_email"] != datos.usuario_id:  # para invitados se usa email como identificador
                raise HTTPException(status_code=403, detail="Solo el usuario que creó la reserva puede cancelarla")

        mp_payment_id = reserva["mp_payment_id"]
        monto = reserva["monto"]
        devolucion_aplicada = "sin_pago"

        # Determinar política de devolución del tenant
        politica = tenant.get("politica_devolucion", "credito")

        if mp_payment_id and monto and float(monto) > 0:
            if politica == "reembolso":
                # Reembolso directo a Mercado Pago
                try:
                    async with httpx.AsyncClient() as client:
                        res = await client.post(
                            f"https://api.mercadopago.com/v1/payments/{mp_payment_id}/refunds",
                            headers={
                                "Authorization": f"Bearer {tenant['mp_access_token']}",
                                "Content-Type": "application/json"
                            },
                            json={}
                        )
                    if res.status_code in (200, 201):
                        await db.execute(
                            "UPDATE reservas SET estado='reembolsada' WHERE id=$1",
                            reserva_id
                        )
                        devolucion_aplicada = "reembolso"
                    else:
                        # Si falla el reembolso, igual cancelamos y otorgamos crédito
                        politica = "credito"
                except Exception:
                    politica = "credito"

            if politica == "credito":
                # Crédito interno para próxima reserva
                await db.execute(
                    """INSERT INTO creditos_reserva
                       (tenant_id, usuario_id, invitado_email, monto, motivo, reserva_origen_id)
                       VALUES ($1,$2,$3,$4,'Cancelación de reserva',$5)""",
                    tenant["id"],
                    reserva["usuario_id"],
                    reserva["invitado_email"],
                    float(monto),
                    reserva_id
                )
                await db.execute(
                    "UPDATE reservas SET estado='cancelada' WHERE id=$1",
                    reserva_id
                )
                devolucion_aplicada = "credito"
        else:
            # Sin pago registrado, cancelación simple
            await db.execute(
                "UPDATE reservas SET estado='cancelada' WHERE id=$1",
                reserva_id
            )

        # Email de cancelación
        email_destino = reserva["invitado_email"]
        nombre_destino = reserva["invitado_nombre"] or "Cliente"
        if not email_destino and reserva["usuario_id"]:
            usuario = await db.fetchrow(
                "SELECT email, nombre FROM usuarios WHERE id=$1",
                reserva["usuario_id"]
            )
            if usuario:
                email_destino = usuario["email"]
                nombre_destino = usuario["nombre"]

        if email_destino:
            msg_devolucion = ""
            if devolucion_aplicada == "reembolso":
                msg_devolucion = "<p style='color:#1A6B3C'><b>💰 El reembolso fue procesado.</b> El dinero volverá a tu medio de pago en 24-72 horas hábiles.</p>"
            elif devolucion_aplicada == "credito":
                msg_devolucion = f"<p style='color:#1A6B3C'><b>💳 Se generó un crédito de ${monto}</b> que podés usar en tu próxima reserva (válido por 6 meses).</p>"

            enviar_email(
                email_destino,
                "❌ Reserva cancelada",
                f"""
                <div style="font-family:Arial,sans-serif;max-width:500px;margin:auto">
                    <h2 style="color:#2C3E50">Reserva cancelada</h2>
                    <p>Hola <b>{nombre_destino}</b>, tu reserva fue cancelada.</p>
                    <table style="border-collapse:collapse;width:100%;margin-top:16px">
                        <tr><td style="padding:10px 14px;font-weight:bold;color:#2C3E50">Fecha</td>
                            <td style="padding:10px 14px">{reserva['fecha']}</td></tr>
                        <tr><td style="padding:10px 14px;font-weight:bold;color:#2C3E50">Horario</td>
                            <td style="padding:10px 14px">{reserva['hora_inicio']} – {reserva['hora_fin']}</td></tr>
                    </table>
                    {msg_devolucion}
                    <p style="margin-top:15px;color:#888">Si no realizaste esta cancelación, contactá al administrador.</p>
                    <p style="color:#888">{tenant['nombre']}</p>
                    <p style="margin-top:20px;font-size:13px;color:#888;text-align:center">
                        <a href="https://reservatuespacio.com/faq-usuarios.html" style="color:#2C3E50;font-weight:bold">Preguntas frecuentes →</a>
                        &nbsp;&nbsp;|&nbsp;&nbsp;
                        <a href="https://reservatuespacio.com/soporte.html" style="color:#888">Centro de soporte →</a>
                    </p>
                </div>
                """
            )

        return {"mensaje": "Reserva cancelada correctamente", "devolucion": devolucion_aplicada}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await release_db(db)

@app.delete("/reservas/{reserva_id}/admin")
async def cancelar_reserva_admin(request: Request, reserva_id: str):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reserva = await db.fetchrow(
            "SELECT id FROM reservas WHERE id=$1 AND estado IN ('activa','pending_payment') AND tenant_id=$2",
            reserva_id, tenant["id"]
        )
        if not reserva:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")
        await db.execute(
            "UPDATE reservas SET estado='cancelada' WHERE id=$1",
            reserva_id
        )
        return {"mensaje": "Reserva cancelada correctamente"}
    finally:
        await release_db(db)

@app.post("/reservas/comprobante")
async def subir_comprobante(request: Request):
    import base64
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        body = await request.json()
        archivo_b64 = body.get("archivo")
        mime = body.get("mime", "image/jpeg")
        reserva_id = body.get("reserva_id")

        if not archivo_b64:
            raise HTTPException(status_code=400, detail="Se requiere el campo archivo en base64")

        import httpx
        archivo_bytes = base64.b64decode(archivo_b64)
        extension = mime.split("/")[-1].replace("jpeg", "jpg")
        import time as time_module
        filename = f"{tenant['slug']}/{int(time_module.time() * 1000)}.{extension}"

        supabase_url = f"https://okkwfaouqdnbityotnje.supabase.co/storage/v1/object/comprobantes/{filename}"
        supabase_key = os.getenv("SUPABASE_SERVICE_KEY", "")
        if not supabase_key:
            raise HTTPException(status_code=500, detail="SUPABASE_SERVICE_KEY no configurada")

        async with httpx.AsyncClient() as client:
            res = await client.put(
                supabase_url,
                content=archivo_bytes,
                headers={"Authorization": f"Bearer {supabase_key}", "Content-Type": mime, "x-upsert": "true"}
            )
        if res.status_code not in (200, 201):
            raise HTTPException(status_code=500, detail=f"Error al subir comprobante: {res.text}")

        url_publica = f"https://okkwfaouqdnbityotnje.supabase.co/storage/v1/object/public/comprobantes/{filename}"

        if reserva_id:
            reserva = await db.fetchrow(
                """SELECT r.*, a.nombre as espacio_nombre,
                          COALESCE(u.nombre, r.invitado_nombre) as cliente_nombre
                   FROM reservas r
                   JOIN aulas a ON r.aula_id = a.id
                   LEFT JOIN usuarios u ON r.usuario_id = u.id
                   WHERE r.id = $1 AND r.tenant_id = $2""",
                reserva_id, tenant["id"]
            )
            if reserva:
                await db.execute("UPDATE reservas SET expires_at = NULL WHERE id = $1", reserva_id)
                if tenant.get("email_admin"):
                    admin_url = f"https://reservatuespacio.com/admin/{tenant['slug']}"
                    enviar_email(
                        tenant["email_admin"],
                        "🧾 Nueva reserva con comprobante de transferencia",
                        f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:auto">
                            <h2 style="color:#2C3E50">Nuevo comprobante recibido</h2>
                            <p>Se realizó una reserva por transferencia y el cliente subió el comprobante. Revisalo y confirmá desde tu panel.</p>
                            <table style="border-collapse:collapse;width:100%;background:#f9f9fb;border-radius:8px">
                                <tr><td style="padding:10px 14px;font-weight:bold">Cliente</td><td style="padding:10px 14px">{reserva['cliente_nombre'] or '—'}</td></tr>
                                <tr><td style="padding:10px 14px;font-weight:bold">Espacio</td><td style="padding:10px 14px">{reserva['espacio_nombre']}</td></tr>
                                <tr><td style="padding:10px 14px;font-weight:bold">Fecha</td><td style="padding:10px 14px">{reserva['fecha']}</td></tr>
                                <tr><td style="padding:10px 14px;font-weight:bold">Horario</td><td style="padding:10px 14px">{reserva['hora_inicio'].strftime('%H:%M')} – {reserva['hora_fin'].strftime('%H:%M')}</td></tr>
                            </table>
                            <div style="margin-top:24px;text-align:center">
                                <a href="{url_publica}" target="_blank" style="display:inline-block;padding:10px 20px;background:#f0f4f8;border:1px solid #ccc;border-radius:8px;color:#2C3E50;text-decoration:none;font-weight:bold;margin-right:10px">📎 Ver comprobante</a>
                                <a href="{admin_url}" style="display:inline-block;padding:10px 20px;background:#2C3E50;border-radius:8px;color:white;text-decoration:none;font-weight:bold">✅ Ir al panel →</a>
                            </div>
                            <p style="margin-top:24px;font-size:12px;color:#aaa;text-align:center">La reserva permanecerá en <b>Pago pendiente</b> hasta que la confirmes o rechaces desde el panel.</p>
                        </div>"""
                    )
        return {"url": url_publica}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


@app.patch("/reservas/{reserva_id}/confirmar-pago")
async def confirmar_pago_transferencia(request: Request, reserva_id: str):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reserva = await db.fetchrow(
            """SELECT r.*, a.nombre as espacio_nombre,
                      COALESCE(u.email, r.invitado_email) as email_destino,
                      COALESCE(u.nombre, r.invitado_nombre) as nombre_destino
               FROM reservas r
               JOIN aulas a ON r.aula_id = a.id
               LEFT JOIN usuarios u ON r.usuario_id = u.id
               WHERE r.id = $1 AND r.estado = 'pending_payment' AND r.tenant_id = $2""",
            reserva_id, tenant["id"]
        )
        if not reserva:
            raise HTTPException(status_code=404, detail="Reserva no encontrada o no está en estado pendiente")
        await db.execute("UPDATE reservas SET estado='activa', expires_at=NULL WHERE id=$1", reserva_id)
        if reserva["email_destino"]:
            enviar_email(
                reserva["email_destino"],
                "✅ Tu reserva fue confirmada",
                f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:auto">
                    <h2 style="color:#2C3E50">¡Reserva confirmada!</h2>
                    <p>Hola <b>{reserva['nombre_destino']}</b>, tu pago fue verificado y tu reserva quedó confirmada.</p>
                    <table style="border-collapse:collapse;width:100%;background:#f9f9fb;border-radius:8px">
                        <tr><td style="padding:10px 14px;font-weight:bold">Espacio</td><td style="padding:10px 14px">{reserva['espacio_nombre']}</td></tr>
                        <tr><td style="padding:10px 14px;font-weight:bold">Fecha</td><td style="padding:10px 14px">{reserva['fecha']}</td></tr>
                        <tr><td style="padding:10px 14px;font-weight:bold">Horario</td><td style="padding:10px 14px">{reserva['hora_inicio'].strftime('%H:%M')} – {reserva['hora_fin'].strftime('%H:%M')}</td></tr>
                    </table>
                </div>"""
            )
        return {"mensaje": "Reserva confirmada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await release_db(db)


@app.get("/usuarios/buscar")
async def buscar_usuario(request: Request, email: str):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        usuario = await db.fetchrow(
            "SELECT * FROM usuarios WHERE email = $1 AND tenant_id = $2",
            email, tenant["id"]
        )
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        return dict(usuario)
    finally:
        await release_db(db)

@app.post("/usuarios")
async def crear_usuario(request: Request, usuario: dict):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        result = await db.fetchrow(
            """INSERT INTO usuarios (email, nombre, rol, tenant_id, whatsapp)
               VALUES ($1, $2, $3, $4, $5) RETURNING id""",
            usuario["email"], usuario["nombre"], usuario.get("rol", "usuario"), tenant["id"],
            usuario.get("whatsapp")
        )
        return {"id": str(result["id"])}
    finally:
        await release_db(db)

@app.get("/usuarios")
async def listar_usuarios(request: Request):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        usuarios = await db.fetch(
            "SELECT * FROM usuarios WHERE tenant_id = $1 ORDER BY nombre",
            tenant["id"]
        )
        return [dict(u) for u in usuarios]
    finally:
        await release_db(db)

@app.patch("/usuarios/{usuario_id}")
async def toggle_usuario(request: Request, usuario_id: str, datos: dict):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        await db.execute(
            "UPDATE usuarios SET activo = $1 WHERE id = $2 AND tenant_id = $3",
            datos["activo"], usuario_id, tenant["id"]
        )
        return {"mensaje": "Usuario actualizado"}
    finally:
        await release_db(db)

@app.get("/reservas/usuario")
async def reservas_por_usuario(request: Request, email: str):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reservas = await db.fetch(
            """SELECT r.id, r.fecha, r.hora_inicio, r.hora_fin, r.estado,
                      r.usuario_id, a.nombre as espacio_nombre,
                      r.invitado_nombre, r.invitado_email
               FROM reservas r
               JOIN aulas a ON r.aula_id = a.id
               LEFT JOIN usuarios u ON r.usuario_id = u.id
               WHERE (u.email = $1 OR r.invitado_email = $1)
               AND r.tenant_id = $2
               ORDER BY r.fecha DESC, r.hora_inicio DESC""",
            email, tenant["id"]
        )
        return [dict(r) for r in reservas]
    finally:
        await release_db(db)

@app.get("/reservas")
async def listar_todas_reservas(request: Request, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        if not fecha_desde:
            fecha_desde = (datetime.utcnow() - timedelta(days=20)).date()
        if not fecha_hasta:
            fecha_hasta = (datetime.utcnow() + timedelta(days=365)).date()
        reservas = await db.fetch(
            """SELECT r.id, r.fecha, r.hora_inicio, r.hora_fin, r.estado,
                      r.usuario_id, a.nombre as espacio_nombre,
                      COALESCE(u.nombre, r.invitado_nombre) as usuario_nombre,
                      COALESCE(u.email, r.invitado_email) as usuario_email,
                      r.invitado_whatsapp, r.observaciones, r.monto,
                      r.comprobante_url
               FROM reservas r
               JOIN aulas a ON r.aula_id = a.id
               LEFT JOIN usuarios u ON r.usuario_id = u.id
               WHERE r.tenant_id = $1
                 AND r.fecha >= $2
                 AND r.fecha <= $3
               ORDER BY r.fecha DESC, r.hora_inicio DESC""",
            tenant["id"], fecha_desde, fecha_hasta
        )
        return [dict(r) for r in reservas]
    finally:
        await release_db(db)

@app.get("/reservas/calendario")
async def reservas_calendario(request: Request, fecha_inicio: date, fecha_fin: date):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reservas = await db.fetch(
            """SELECT r.id, r.fecha, r.hora_inicio, r.hora_fin,
                      a.id as espacio_id, a.nombre as espacio_nombre,
                      COALESCE(u.nombre, r.invitado_nombre) as usuario_nombre
               FROM reservas r
               JOIN aulas a ON r.aula_id = a.id
               LEFT JOIN usuarios u ON r.usuario_id = u.id
               WHERE r.fecha BETWEEN $1 AND $2
               AND r.estado = 'activa'
               AND r.tenant_id = $3
               ORDER BY r.fecha, r.hora_inicio""",
            fecha_inicio, fecha_fin, tenant["id"]
        )
        return [dict(r) for r in reservas]
    finally:
        await release_db(db)

@app.get("/fechas-bloqueadas")
async def listar_fechas_bloqueadas(request: Request):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        fechas = await db.fetch(
            "SELECT * FROM fechas_bloqueadas WHERE tenant_id = $1 ORDER BY fecha",
            tenant["id"]
        )
        return [dict(f) for f in fechas]
    finally:
        await release_db(db)

@app.post("/fechas-bloqueadas")
async def agregar_fecha_bloqueada(request: Request, datos: FechaBloqueada):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        await db.execute(
            "INSERT INTO fechas_bloqueadas (fecha, motivo, tenant_id) VALUES ($1, $2, $3)",
            datos.fecha, datos.motivo, tenant["id"]
        )
        return {"mensaje": f"Fecha {datos.fecha} bloqueada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.delete("/fechas-bloqueadas/{fecha}")
async def eliminar_fecha_bloqueada(request: Request, fecha: date):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        result = await db.execute(
            "DELETE FROM fechas_bloqueadas WHERE fecha = $1 AND tenant_id = $2",
            fecha, tenant["id"]
        )
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="Fecha no encontrada")
        return {"mensaje": "Fecha desbloqueada correctamente"}
    finally:
        await release_db(db)

@app.get("/horarios")
async def listar_horarios(request: Request):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        horarios = await db.fetch(
            "SELECT * FROM configuracion_horarios WHERE tenant_id = $1 ORDER BY dia_semana",
            tenant["id"]
        )
        return [dict(h) for h in horarios]
    finally:
        await release_db(db)

@app.patch("/horarios/{dia_semana}")
async def actualizar_horario(request: Request, dia_semana: int, datos: HorarioUpdate):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        await db.execute(
            """UPDATE configuracion_horarios
               SET habilitado = $1, hora_apertura = $2, hora_cierre = $3
               WHERE dia_semana = $4 AND tenant_id = $5""",
            datos.habilitado, datos.hora_apertura, datos.hora_cierre, dia_semana, tenant["id"]
        )
        return {"mensaje": "Horario actualizado"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


# ============================================================
# TURNOS FIJOS — CRUD para admin + endpoint público
# ============================================================

class TurnoFijoCreate(BaseModel):
    aula_id: str
    dia_semana: Optional[int] = None       # None si es puntual con fecha_especifica
    hora_inicio: str
    hora_fin: str
    activo: bool = True
    fecha_especifica: Optional[date] = None  # Si tiene fecha, es turno puntual (no recurrente)
    es_excepcion: bool = False               # True = anulación de turno recurrente ese día

class TurnoFijoUpdate(BaseModel):
    hora_inicio: Optional[str] = None
    hora_fin: Optional[str] = None
    activo: Optional[bool] = None
    es_excepcion: Optional[bool] = None


@app.get("/turnos-fijos")
async def listar_turnos_fijos_publico(request: Request, aula_id: Optional[str] = None):
    """Devuelve los turnos fijos activos del tenant, opcionalmente filtrados por espacio."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        if aula_id:
            turnos = await db.fetch(
                """SELECT id, aula_id::text, dia_semana, hora_inicio::text, hora_fin::text, activo
                   FROM turnos_fijos
                   WHERE tenant_id = $1 AND aula_id = $2 AND activo = TRUE
                   ORDER BY dia_semana, hora_inicio""",
                tenant["id"], aula_id
            )
        else:
            turnos = await db.fetch(
                """SELECT id, aula_id::text, dia_semana, hora_inicio::text, hora_fin::text, activo
                   FROM turnos_fijos
                   WHERE tenant_id = $1 AND activo = TRUE
                   ORDER BY dia_semana, hora_inicio""",
                tenant["id"]
            )
        return [dict(t) for t in turnos]
    finally:
        await release_db(db)


@app.get("/turnos-fijos/disponibilidad")
async def turnos_disponibilidad(request: Request, fecha: str, aula_id: Optional[str] = None):
    """
    Devuelve los turnos del día con estado: libre o reservado.
    Combina turnos recurrentes (por dia_semana) y puntuales (fecha_especifica).
    Aplica excepciones (es_excepcion=True) que anulan turnos recurrentes.
    """
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        fecha_obj = date.fromisoformat(fecha)
        dia_semana = fecha_obj.weekday()
        dia_js = (dia_semana + 1) % 7  # Python 0=lunes → JS 1=lunes, 0=domingo

        aula_filtro = f"AND tf.aula_id = '{aula_id}'" if aula_id else ""

        # Obtener todos los turnos relevantes para este día:
        # 1. Recurrentes que coinciden con el día de la semana (sin fecha_especifica)
        # 2. Puntuales para esta fecha exacta
        # 3. Excepciones para esta fecha exacta (para anular recurrentes)
        turnos_raw = await db.fetch(
            f"""SELECT tf.id, tf.aula_id::text, tf.dia_semana,
                       tf.hora_inicio::text, tf.hora_fin::text,
                       tf.activo, tf.fecha_especifica::text, tf.es_excepcion
                FROM turnos_fijos tf
                WHERE tf.tenant_id = $1
                  AND tf.activo = TRUE
                  {aula_filtro}
                  AND (
                      -- Recurrentes del día de la semana
                      (tf.fecha_especifica IS NULL AND tf.dia_semana = $2 AND tf.es_excepcion = FALSE)
                      OR
                      -- Puntuales para esta fecha
                      (tf.fecha_especifica = $3 AND tf.es_excepcion = FALSE)
                      OR
                      -- Excepciones para esta fecha (anulaciones)
                      (tf.fecha_especifica = $3 AND tf.es_excepcion = TRUE)
                  )
                ORDER BY tf.hora_inicio""",
            tenant["id"], dia_js, fecha_obj
        )

        # Separar excepciones de turnos normales
        excepciones_horas = set()
        turnos_normales = []
        for t in turnos_raw:
            if t["es_excepcion"]:
                excepciones_horas.add(t["hora_inicio"][:5])
            else:
                turnos_normales.append(t)

        # Filtrar turnos recurrentes que tienen excepción este día
        turnos = [t for t in turnos_normales if t["hora_inicio"][:5] not in excepciones_horas]

        # Reservas del día para este espacio
        if aula_id:
            reservas = await db.fetch(
                """SELECT r.hora_inicio::text, r.hora_fin::text,
                          COALESCE(u.nombre, r.invitado_nombre) as cliente_nombre,
                          COALESCE(u.email, r.invitado_email) as cliente_email,
                          r.estado, r.id as reserva_id, a.nombre as espacio_nombre
                   FROM reservas r
                   LEFT JOIN usuarios u ON r.usuario_id = u.id
                   LEFT JOIN aulas a ON r.aula_id = a.id
                   WHERE r.tenant_id = $1 AND r.aula_id = $2 AND r.fecha = $3
                   AND r.estado NOT IN ('cancelada', 'expirada')
                   ORDER BY r.hora_inicio""",
                tenant["id"], aula_id, fecha_obj
            )
        else:
            reservas = await db.fetch(
                """SELECT r.hora_inicio::text, r.hora_fin::text,
                          COALESCE(u.nombre, r.invitado_nombre) as cliente_nombre,
                          COALESCE(u.email, r.invitado_email) as cliente_email,
                          r.estado, r.id as reserva_id, a.nombre as espacio_nombre
                   FROM reservas r
                   LEFT JOIN usuarios u ON r.usuario_id = u.id
                   LEFT JOIN aulas a ON r.aula_id = a.id
                   WHERE r.tenant_id = $1 AND r.fecha = $2
                   AND r.estado NOT IN ('cancelada', 'expirada')
                   ORDER BY r.hora_inicio""",
                tenant["id"], fecha_obj
            )

        reservas_dict = [dict(r) for r in reservas]

        resultado_turnos = []
        for t in turnos:
            turno_ocupado = next(
                (r for r in reservas_dict if r["hora_inicio"][:5] == t["hora_inicio"][:5]),
                None
            )
            resultado_turnos.append({
                "id": str(t["id"]),
                "aula_id": str(t["aula_id"]) if t["aula_id"] else None,
                "hora_inicio": t["hora_inicio"][:5],
                "hora_fin": t["hora_fin"][:5],
                "libre": turno_ocupado is None,
                "es_puntual": t["fecha_especifica"] is not None,
                "reserva": turno_ocupado
            })

        return {
            "fecha": fecha,
            "dia_semana": dia_js,
            "turnos_fijos": resultado_turnos,
            "reservas_libres": reservas_dict
        }
    finally:
        await release_db(db)


@app.get("/admin/turnos-fijos")
async def admin_listar_turnos_fijos(request: Request):
    """Lista todos los turnos fijos del tenant para el panel admin, con nombre del espacio."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        turnos = await db.fetch(
            """SELECT tf.id, tf.aula_id::text, tf.dia_semana,
                      tf.hora_inicio::text, tf.hora_fin::text, tf.activo,
                      tf.fecha_especifica::text, tf.es_excepcion,
                      a.nombre as espacio_nombre
               FROM turnos_fijos tf
               LEFT JOIN aulas a ON tf.aula_id = a.id
               WHERE tf.tenant_id = $1
               ORDER BY a.nombre, tf.fecha_especifica NULLS LAST, tf.dia_semana, tf.hora_inicio""",
            tenant["id"]
        )
        return [dict(t) for t in turnos]
    finally:
        await release_db(db)


@app.post("/admin/turnos-fijos")
async def admin_crear_turno_fijo(request: Request, datos: TurnoFijoCreate):
    """Crea un turno fijo para un espacio específico. Puede ser recurrente, puntual o excepción."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        from datetime import time as time_type
        def parse_time(t):
            h, m = map(int, t.split(':')[:2])
            return time_type(h, m)

        # Validar: necesita día_semana (recurrente) o fecha_especifica (puntual)
        if datos.fecha_especifica is None and datos.dia_semana is None:
            raise HTTPException(status_code=400, detail="Indicá día de la semana o fecha específica")

        # Verificar que el espacio pertenece al tenant
        espacio = await db.fetchrow(
            "SELECT id, nombre FROM aulas WHERE id = $1 AND tenant_id = $2",
            datos.aula_id, tenant["id"]
        )
        if not espacio:
            raise HTTPException(status_code=404, detail="Espacio no encontrado")

        # Para turnos recurrentes: si no viene dia_semana pero sí fecha, calcular el dia_semana
        dia_semana = datos.dia_semana
        if dia_semana is None and datos.fecha_especifica:
            # Python weekday(): 0=lunes. Nuestra convención JS: 0=domingo, 1=lunes...
            dia_semana = (datos.fecha_especifica.weekday() + 1) % 7

        turno = await db.fetchrow(
            """INSERT INTO turnos_fijos
               (tenant_id, aula_id, dia_semana, hora_inicio, hora_fin, activo, fecha_especifica, es_excepcion)
               VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
               RETURNING id, aula_id::text, dia_semana, hora_inicio::text, hora_fin::text,
                         activo, fecha_especifica::text, es_excepcion""",
            tenant["id"], datos.aula_id, dia_semana,
            parse_time(datos.hora_inicio), parse_time(datos.hora_fin),
            datos.activo, datos.fecha_especifica, datos.es_excepcion
        )
        result = dict(turno)
        result["espacio_nombre"] = espacio["nombre"]
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


@app.patch("/admin/turnos-fijos/{turno_id}")
async def admin_actualizar_turno_fijo(request: Request, turno_id: str, datos: TurnoFijoUpdate):
    """Actualiza un turno fijo."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        from datetime import time as time_type
        def parse_time(t):
            h, m = map(int, t.split(':')[:2])
            return time_type(h, m)

        campos = {k: v for k, v in datos.dict().items() if v is not None}
        if not campos:
            raise HTTPException(status_code=400, detail="No hay campos para actualizar")
        # Convertir strings de hora a time
        if 'hora_inicio' in campos and isinstance(campos['hora_inicio'], str):
            campos['hora_inicio'] = parse_time(campos['hora_inicio'])
        if 'hora_fin' in campos and isinstance(campos['hora_fin'], str):
            campos['hora_fin'] = parse_time(campos['hora_fin'])
        sets = ", ".join([f"{k} = ${i+3}" for i, k in enumerate(campos.keys())])
        turno = await db.fetchrow(
            f"""UPDATE turnos_fijos SET {sets}
                WHERE id = $1 AND tenant_id = $2
                RETURNING id, dia_semana, hora_inicio::text, hora_fin::text, activo""",
            turno_id, tenant["id"], *list(campos.values())
        )
        if not turno:
            raise HTTPException(status_code=404, detail="Turno no encontrado")
        return dict(turno)
    finally:
        await release_db(db)


@app.delete("/admin/turnos-fijos/{turno_id}")
async def admin_eliminar_turno_fijo(request: Request, turno_id: str):
    """Elimina un turno fijo."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        await db.execute(
            "DELETE FROM turnos_fijos WHERE id = $1 AND tenant_id = $2",
            turno_id, tenant["id"]
        )
        return {"ok": True}
    finally:
        await release_db(db)


@app.get("/admin/agenda")
async def admin_agenda(request: Request, fecha_inicio: str, fecha_fin: str):
    """
    Panel visual semanal/mensual para el admin.
    Devuelve todos los turnos del período con info del cliente y estado.
    """
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reservas = await db.fetch(
            """SELECT r.id, r.fecha::text, r.hora_inicio::text, r.hora_fin::text,
                      r.estado, r.observaciones,
                      COALESCE(u.nombre, r.invitado_nombre) as cliente_nombre,
                      COALESCE(u.email, r.invitado_email) as cliente_email,
                      COALESCE(u.whatsapp, r.invitado_whatsapp) as cliente_whatsapp,
                      a.nombre as espacio_nombre,
                      r.monto, r.mp_payment_id
               FROM reservas r
               LEFT JOIN usuarios u ON r.usuario_id = u.id
               LEFT JOIN aulas a ON r.aula_id = a.id
               WHERE r.tenant_id = $1
               AND r.fecha BETWEEN $2 AND $3
               AND r.estado NOT IN ('cancelada', 'expirada')
               ORDER BY r.fecha, r.hora_inicio""",
            tenant["id"], fecha_inicio, fecha_fin
        )
        return [dict(r) for r in reservas]
    finally:
        await release_db(db)


async def listar_edificios(request: Request):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        edificios = await db.fetch(
            "SELECT * FROM edificios WHERE activo = TRUE AND tenant_id = $1",
            tenant["id"]
        )
        return [dict(e) for e in edificios]
    finally:
        await release_db(db)

@app.post("/edificios")
async def crear_edificio(request: Request, datos: dict):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        result = await db.fetchrow(
            "INSERT INTO edificios (nombre, direccion, tenant_id) VALUES ($1, $2, $3) RETURNING id",
            datos["nombre"], datos.get("direccion", ""), tenant["id"]
        )
        return {"mensaje": "Edificio creado", "id": result["id"]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.patch("/edificios/{edificio_id}")
async def actualizar_edificio(request: Request, edificio_id: int, datos: dict):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        if "nombre" in datos:
            await db.execute(
                "UPDATE edificios SET nombre = $1 WHERE id = $2 AND tenant_id = $3",
                datos["nombre"], edificio_id, tenant["id"]
            )
        if "activo" in datos:
            await db.execute(
                "UPDATE edificios SET activo = $1 WHERE id = $2 AND tenant_id = $3",
                datos["activo"], edificio_id, tenant["id"]
            )
        return {"mensaje": "Edificio actualizado"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.get("/edificios/{edificio_id}/espacios")
async def espacios_por_edificio(request: Request, edificio_id: int):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        espacios = await db.fetch(
            "SELECT * FROM aulas WHERE edificio_id = $1 AND activa = TRUE AND tenant_id = $2",
            edificio_id, tenant["id"]
        )
        return [dict(e) for e in espacios]
    finally:
        await release_db(db)

@app.get("/reservas/exportar")
async def exportar_reservas(request: Request, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None):
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        filtros = [tenant["id"]]
        where_extra = ""
        if fecha_desde:
            filtros.append(fecha_desde)
            where_extra += f" AND r.fecha >= ${len(filtros)}"
        if fecha_hasta:
            filtros.append(fecha_hasta)
            where_extra += f" AND r.fecha <= ${len(filtros)}"
        reservas = await db.fetch(
            f"""SELECT r.fecha, r.hora_inicio, r.hora_fin, r.estado,
                      a.nombre as espacio_nombre,
                      COALESCE(u.nombre, r.invitado_nombre) as usuario_nombre,
                      COALESCE(u.email, r.invitado_email) as usuario_email,
                      COALESCE(r.invitado_whatsapp, '') as whatsapp,
                      COALESCE(r.monto::text, '') as monto
               FROM reservas r
               JOIN aulas a ON r.aula_id = a.id
               LEFT JOIN usuarios u ON r.usuario_id = u.id
               WHERE r.tenant_id = $1{where_extra}
               ORDER BY r.fecha DESC, r.hora_inicio DESC""",
            *filtros
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"
        encabezados = ["Fecha", "Hora inicio", "Hora fin", "Espacio", "Usuario", "Email", "WhatsApp", "Monto", "Estado"]
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=enc)
            celda.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            celda.fill = openpyxl.styles.PatternFill("solid", fgColor="1B4F8A")
        for fila, r in enumerate(reservas, 2):
            ws.cell(row=fila, column=1, value=r["fecha"].strftime("%d/%m/%Y"))
            ws.cell(row=fila, column=2, value=r["hora_inicio"].strftime("%H:%M"))
            ws.cell(row=fila, column=3, value=r["hora_fin"].strftime("%H:%M"))
            ws.cell(row=fila, column=4, value=r["espacio_nombre"])
            ws.cell(row=fila, column=5, value=r["usuario_nombre"] or "—")
            ws.cell(row=fila, column=6, value=r["usuario_email"] or "—")
            ws.cell(row=fila, column=7, value=r["whatsapp"] or "—")
            ws.cell(row=fila, column=8, value=r["monto"] or "—")
            ws.cell(row=fila, column=9, value=r["estado"])
        anchos = [12, 12, 12, 20, 25, 30, 16, 12, 15]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
        desde_str = fecha_desde.strftime("%Y%m%d") if fecha_desde else "inicio"
        hasta_str = fecha_hasta.strftime("%Y%m%d") if fecha_hasta else "hoy"
        filename = f"reservas_{desde_str}_{hasta_str}.xlsx"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        await release_db(db)



# ===== RECURSOS =====

class RecursoCreate(BaseModel):
    nombre: str
    cantidad_total: int

class RecursoUpdate(BaseModel):
    nombre: Optional[str] = None
    cantidad_total: Optional[int] = None
    activo: Optional[bool] = None

@app.get("/recursos")
async def listar_recursos(request: Request):
    """Lista los recursos activos del tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        recursos = await db.fetch(
            "SELECT * FROM recursos WHERE tenant_id = $1 AND activo = TRUE ORDER BY nombre",
            tenant["id"]
        )
        return [dict(r) for r in recursos]
    finally:
        await release_db(db)

@app.get("/recursos/disponibilidad")
async def disponibilidad_recursos(request: Request, fecha: date, hora_inicio: time, hora_fin: time):
    """Devuelve los recursos con su disponibilidad para un horario dado."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        recursos = await db.fetch(
            "SELECT * FROM recursos WHERE tenant_id = $1 AND activo = TRUE ORDER BY nombre",
            tenant["id"]
        )
        resultado = []
        for r in recursos:
            en_uso = await db.fetchval(
                """SELECT COALESCE(SUM(rr.cantidad), 0)
                   FROM recursos_reservas rr
                   JOIN reservas res ON rr.reserva_id = res.id
                   WHERE rr.recurso_id = $1
                   AND res.fecha = $2
                   AND res.estado = 'activa'
                   AND res.tenant_id = $3
                   AND res.hora_inicio < $4 AND res.hora_fin > $5""",
                r["id"], fecha, tenant["id"], hora_fin, hora_inicio
            )
            disponibles = r["cantidad_total"] - int(en_uso)
            resultado.append({
                **dict(r),
                "disponibles": disponibles
            })
        return resultado
    finally:
        await release_db(db)

@app.post("/admin/recursos")
async def crear_recurso(request: Request, datos: RecursoCreate):
    """Crea un nuevo recurso para el tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        result = await db.fetchrow(
            """INSERT INTO recursos (tenant_id, nombre, cantidad_total)
               VALUES ($1, $2, $3) RETURNING *""",
            tenant["id"], datos.nombre, datos.cantidad_total
        )
        return dict(result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.get("/admin/recursos")
async def listar_recursos_admin(request: Request):
    """Lista todos los recursos del tenant (admin)."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        recursos = await db.fetch(
            "SELECT * FROM recursos WHERE tenant_id = $1 ORDER BY nombre",
            tenant["id"]
        )
        return [dict(r) for r in recursos]
    finally:
        await release_db(db)

@app.patch("/admin/recursos/{recurso_id}")
async def actualizar_recurso(request: Request, recurso_id: str, datos: RecursoUpdate):
    """Actualiza un recurso del tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        recurso = await db.fetchrow(
            "SELECT id FROM recursos WHERE id = $1 AND tenant_id = $2",
            recurso_id, tenant["id"]
        )
        if not recurso:
            raise HTTPException(status_code=404, detail="Recurso no encontrado")
        if datos.nombre is not None:
            await db.execute("UPDATE recursos SET nombre = $1 WHERE id = $2", datos.nombre, recurso_id)
        if datos.cantidad_total is not None:
            await db.execute("UPDATE recursos SET cantidad_total = $1 WHERE id = $2", datos.cantidad_total, recurso_id)
        if datos.activo is not None:
            await db.execute("UPDATE recursos SET activo = $1 WHERE id = $2", datos.activo, recurso_id)
        return {"mensaje": "Recurso actualizado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.delete("/admin/recursos/{recurso_id}")
async def eliminar_recurso(request: Request, recurso_id: str):
    """Elimina un recurso del tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        result = await db.execute(
            "DELETE FROM recursos WHERE id = $1 AND tenant_id = $2",
            recurso_id, tenant["id"]
        )
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="Recurso no encontrado")
        return {"mensaje": "Recurso eliminado correctamente"}
    finally:
        await release_db(db)

# ===== OBSERVACIONES EN RESERVAS =====

class ObservacionUpdate(BaseModel):
    observaciones: str

@app.patch("/reservas/{reserva_id}/observaciones")
async def agregar_observacion(request: Request, reserva_id: str, datos: ObservacionUpdate):
    """El admin del tenant agrega observaciones a una reserva."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        reserva = await db.fetchrow(
            "SELECT id FROM reservas WHERE id = $1 AND tenant_id = $2",
            reserva_id, tenant["id"]
        )
        if not reserva:
            raise HTTPException(status_code=404, detail="Reserva no encontrada")
        await db.execute(
            "UPDATE reservas SET observaciones = $1 WHERE id = $2",
            datos.observaciones, reserva_id
        )
        return {"mensaje": "Observación guardada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

# ===== REGISTRO DE NUEVO TENANT =====

class RegistroCreate(BaseModel):
    nombre: str
    tipo: str
    email_admin: str
    nombre_admin: str
    whatsapp_admin: Optional[str] = None
    plan_id: str
    slug: str

@app.post("/registro")
async def registrar_tenant(datos: RegistroCreate):
    db = await get_db()
    try:
        # Validar que el slug no esté tomado
        existente = await db.fetchrow(
            "SELECT id FROM tenants WHERE slug = $1",
            datos.slug
        )
        if existente:
            raise HTTPException(
                status_code=400,
                detail="Esa URL ya está en uso. Por favor elegí otra."
            )

        # Validar que el email no esté registrado
        email_existente = await db.fetchrow(
            "SELECT id FROM tenants WHERE email_admin = $1",
            datos.email_admin
        )
        if email_existente:
            raise HTTPException(
                status_code=400,
                detail="Ya existe una cuenta registrada con ese email."
            )

        # Crear el tenant
        tenant = await db.fetchrow(
            """INSERT INTO tenants (nombre, slug, email_admin, plan_id, trial_hasta, suscripcion_activa)
               VALUES ($1, $2, $3, $4, CURRENT_DATE + INTERVAL '30 days', FALSE)
               RETURNING id, nombre, slug, trial_hasta""",
            datos.nombre, datos.slug, datos.email_admin, datos.plan_id
        )

        tenant_id = tenant["id"]

        # Crear horarios por defecto (lunes a viernes 8:00-20:00, finde cerrado)
        dias = [
            (0, "Lunes",     True,  time(8, 0), time(20, 0)),
            (1, "Martes",    True,  time(8, 0), time(20, 0)),
            (2, "Miércoles", True,  time(8, 0), time(20, 0)),
            (3, "Jueves",    True,  time(8, 0), time(20, 0)),
            (4, "Viernes",   True,  time(8, 0), time(20, 0)),
            (5, "Sábado",    False, None,       None),
            (6, "Domingo",   False, None,       None),
        ]

        for dia_semana, nombre_dia, habilitado, apertura, cierre in dias:
            await db.execute(
                """INSERT INTO configuracion_horarios
                   (dia_semana, nombre_dia, habilitado, hora_apertura, hora_cierre, tenant_id)
                   VALUES ($1, $2, $3, $4, $5, $6)""",
                dia_semana, nombre_dia, habilitado, apertura, cierre, tenant_id
            )

        # Generar contraseña temporal y crear usuario admin
        trial_hasta = tenant["trial_hasta"].strftime("%d/%m/%Y")
        password_temp = generar_password_temporal()
        password_hash = bcrypt.hashpw(password_temp.encode(), bcrypt.gensalt()).decode()
        await db.execute(
            """INSERT INTO usuarios (tenant_id, email, nombre, rol, activo, password_hash, whatsapp)
               VALUES ($1, $2, $3, 'admin', TRUE, $4, $5)""",
            tenant_id, datos.email_admin, datos.nombre_admin, password_hash, datos.whatsapp_admin
        )

        # Enviar email de bienvenida con credenciales
        enviar_email(
            datos.email_admin,
            "¡Bienvenido a ReservaTuEspacio! — Tus datos de acceso",
            email_bienvenida_tenant(
                datos.nombre_admin, datos.nombre, datos.slug,
                datos.email_admin, password_temp, trial_hasta
            )
        )

        return {
            "mensaje": "Cuenta creada exitosamente",
            "tenant_id": str(tenant_id),
            "slug": datos.slug,
            "trial_hasta": trial_hasta
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


# ===== SUPERADMIN =====

def verificar_superadmin(request: Request):
    token = request.headers.get("X-Superadmin-Token", "")
    password_correcta = os.getenv("ADMIN_PASSWORD", "sL2#di!KBw")
    if token != password_correcta:
        raise HTTPException(status_code=401, detail="No autorizado")

class TenantUpdate(BaseModel):
    activo: Optional[bool] = None
    suscripcion_activa: Optional[bool] = None
    plan_id: Optional[str] = None
    trial_hasta: Optional[date] = None

class PlanUpdate(BaseModel):
    nombre: Optional[str] = None
    precio_mensual: Optional[float] = None
    descripcion: Optional[str] = None

class TenantCreate(BaseModel):
    nombre: str
    slug: str
    email_admin: str
    nombre_admin: str
    plan_id: str
    trial_dias: int = 30

@app.get("/superadmin/tenants")
async def superadmin_listar_tenants(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        tenants = await db.fetch("""
            SELECT
                t.*,
                p.nombre as plan_nombre,
                p.precio_mensual as plan_precio,
                (SELECT COUNT(*) FROM usuarios u WHERE u.tenant_id = t.id) as total_usuarios,
                (SELECT COUNT(*) FROM reservas r WHERE r.tenant_id = t.id AND r.estado = 'activa') as total_reservas,
                (SELECT whatsapp FROM usuarios u WHERE u.tenant_id = t.id AND u.rol = 'admin' LIMIT 1) as whatsapp_admin
            FROM tenants t
            LEFT JOIN planes p ON t.plan_id = p.id
            ORDER BY t.activo DESC, t.trial_hasta DESC
        """)
        return [dict(t) for t in tenants]
    finally:
        await release_db(db)

@app.patch("/superadmin/tenants/{tenant_id}")
async def superadmin_actualizar_tenant(request: Request, tenant_id: str, datos: TenantUpdate):
    verificar_superadmin(request)
    db = await get_db()
    try:
        tenant = await db.fetchrow("SELECT id FROM tenants WHERE id = $1", tenant_id)
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant no encontrado")
        if datos.activo is not None:
            await db.execute("UPDATE tenants SET activo = $1 WHERE id = $2", datos.activo, tenant_id)
        if datos.suscripcion_activa is not None:
            if datos.suscripcion_activa:
                # Al activar manualmente: calcular fecha de vencimiento y resetear flags de avisos
                await db.execute(
                    """UPDATE tenants
                       SET suscripcion_activa = TRUE,
                           suscripcion_vence = CURRENT_DATE + INTERVAL '30 days',
                           aviso_7_enviado = FALSE,
                           aviso_3_enviado = FALSE,
                           aviso_1_enviado = FALSE
                       WHERE id = $1""",
                    tenant_id
                )
            else:
                await db.execute("UPDATE tenants SET suscripcion_activa = FALSE WHERE id = $1", tenant_id)
        if datos.plan_id is not None:
            await db.execute("UPDATE tenants SET plan_id = $1 WHERE id = $2", datos.plan_id, tenant_id)
        if datos.trial_hasta is not None:
            await db.execute("UPDATE tenants SET trial_hasta = $1 WHERE id = $2", datos.trial_hasta, tenant_id)
        return {"mensaje": "Tenant actualizado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.post("/superadmin/tenants")
async def superadmin_crear_tenant(request: Request, datos: TenantCreate):
    verificar_superadmin(request)
    db = await get_db()
    try:
        existente = await db.fetchrow("SELECT id FROM tenants WHERE slug = $1", datos.slug)
        if existente:
            raise HTTPException(status_code=400, detail="Ese slug ya está en uso")
        email_existente = await db.fetchrow("SELECT id FROM tenants WHERE email_admin = $1", datos.email_admin)
        if email_existente:
            raise HTTPException(status_code=400, detail="Ya existe una cuenta con ese email")
        tenant = await db.fetchrow(
            """INSERT INTO tenants (nombre, slug, email_admin, plan_id, trial_hasta, suscripcion_activa)
               VALUES ($1, $2, $3, $4, CURRENT_DATE + ($5 || ' days')::interval, FALSE)
               RETURNING id, nombre, slug, trial_hasta""",
            datos.nombre, datos.slug, datos.email_admin, datos.plan_id, str(datos.trial_dias)
        )
        tenant_id = tenant["id"]
        dias = [
            (0, "Lunes",     True,  time(8, 0), time(20, 0)),
            (1, "Martes",    True,  time(8, 0), time(20, 0)),
            (2, "Miércoles", True,  time(8, 0), time(20, 0)),
            (3, "Jueves",    True,  time(8, 0), time(20, 0)),
            (4, "Viernes",   True,  time(8, 0), time(20, 0)),
            (5, "Sábado",    False, None,        None),
            (6, "Domingo",   False, None,        None),
        ]
        for dia_semana, nombre_dia, habilitado, apertura, cierre in dias:
            await db.execute(
                """INSERT INTO configuracion_horarios
                   (dia_semana, nombre_dia, habilitado, hora_apertura, hora_cierre, tenant_id)
                   VALUES ($1, $2, $3, $4, $5, $6)""",
                dia_semana, nombre_dia, habilitado, apertura, cierre, tenant_id
            )
        trial_hasta = tenant["trial_hasta"].strftime("%d/%m/%Y")

        # Generar contraseña temporal y crear usuario admin
        password_temp = generar_password_temporal()
        password_hash = bcrypt.hashpw(password_temp.encode(), bcrypt.gensalt()).decode()
        await db.execute(
            """INSERT INTO usuarios (tenant_id, email, nombre, rol, activo, password_hash, whatsapp)
               VALUES ($1, $2, $3, 'admin', TRUE, $4, $5)""",
            tenant_id, datos.email_admin, datos.nombre_admin, password_hash, datos.whatsapp_admin
        )

        # Enviar email de bienvenida con credenciales
        enviar_email(
            datos.email_admin,
            "¡Bienvenido a ReservaTuEspacio! — Tus datos de acceso",
            email_bienvenida_tenant(
                datos.nombre_admin, datos.nombre, datos.slug,
                datos.email_admin, password_temp, trial_hasta
            )
        )
        # Registrar envío en emails_bienvenida
        try:
            await db.execute(
                """INSERT INTO emails_bienvenida (tenant_id, enviado_a, enviado_por)
                   VALUES ($1, $2, 'automatico')""",
                tenant_id, datos.email_admin
            )
        except Exception:
            pass  # No bloquear si falla el registro
        return {
            "mensaje": "Tenant creado correctamente",
            "tenant_id": str(tenant_id),
            "slug": datos.slug,
            "trial_hasta": trial_hasta
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.delete("/superadmin/tenants/{tenant_id}")
async def superadmin_eliminar_tenant(request: Request, tenant_id: str):
    verificar_superadmin(request)
    db = await get_db()
    try:
        tenant = await db.fetchrow("SELECT id, nombre FROM tenants WHERE id = $1", tenant_id)
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant no encontrado")
        await db.execute("DELETE FROM reservas WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM usuarios WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM aulas WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM edificios WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM configuracion_horarios WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM fechas_bloqueadas WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM pagos WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM tickets WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM recursos WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM mp_preferencias WHERE tenant_id = $1", tenant_id)
        await db.execute("DELETE FROM tenants WHERE id = $1", tenant_id)
        return {"mensaje": f"Tenant '{tenant['nombre']}' eliminado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.patch("/superadmin/tenants/{tenant_id}/whatsapp")
async def superadmin_actualizar_whatsapp(request: Request, tenant_id: str):
    verificar_superadmin(request)
    db = await get_db()
    try:
        body = await request.json()
        whatsapp = body.get('whatsapp')
        tenant = await db.fetchrow('SELECT id FROM tenants WHERE id = $1', tenant_id)
        if not tenant:
            raise HTTPException(status_code=404, detail='Tenant no encontrado')
        # Actualizar whatsapp en el usuario admin del tenant
        await db.execute(
            """UPDATE usuarios SET whatsapp = $1
               WHERE tenant_id = $2 AND rol = 'admin'""",
            whatsapp, tenant_id
        )
        return {'mensaje': 'WhatsApp actualizado correctamente'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await release_db(db)


@app.post("/superadmin/tenants/{tenant_id}/reenviar-bienvenida")
async def superadmin_reenviar_bienvenida(request: Request, tenant_id: str):
    verificar_superadmin(request)
    db = await get_db()
    try:
        # Obtener datos del tenant
        tenant = await db.fetchrow(
            """SELECT t.*, p.nombre as plan_nombre
               FROM tenants t
               LEFT JOIN planes p ON t.plan_id = p.id
               WHERE t.id = $1""",
            tenant_id
        )
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant no encontrado")

        # Obtener admin del tenant
        admin = await db.fetchrow(
            """SELECT nombre, email FROM usuarios
               WHERE tenant_id = $1 AND rol = 'admin'
               LIMIT 1""",
            tenant_id
        )
        if not admin:
            raise HTTPException(status_code=404, detail="No se encontró el usuario admin del tenant")

        slug = tenant['slug']
        trial_hasta = tenant['trial_hasta'].strftime('%d/%m/%Y') if tenant.get('trial_hasta') else 'N/A'

        # Enviar email con contraseña temporal genérica (el admin debe cambiarla)
        html = email_bienvenida_tenant(
            nombre_admin=admin['nombre'],
            nombre_tenant=tenant['nombre'],
            slug=slug,
            email=admin['email'],
            password_temp='(usá la contraseña que te enviamos anteriormente o contactá a soporte)',
            trial_hasta=trial_hasta
        )
        enviar_email(admin['email'], '¡Bienvenido a ReservaTuEspacio! Tus datos de acceso', html)

        # Registrar en emails_bienvenida
        await db.execute(
            """INSERT INTO emails_bienvenida (tenant_id, enviado_a, enviado_por)
               VALUES ($1, $2, 'superadmin')""",
            tenant_id, admin['email']
        )

        return {"mensaje": f"Email de bienvenida reenviado a {admin['email']}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await release_db(db)


@app.get("/superadmin/emails-bienvenida")
async def superadmin_listar_emails_bienvenida(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        emails = await db.fetch(
            """SELECT e.*, t.nombre as tenant_nombre, t.slug
               FROM emails_bienvenida e
               LEFT JOIN tenants t ON e.tenant_id = t.id
               ORDER BY e.created_at DESC
               LIMIT 100"""
        )
        return [dict(e) for e in emails]
    except Exception as ex:
        return []
    finally:
        await release_db(db)


@app.get("/superadmin/planes")
async def superadmin_listar_planes(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        planes = await db.fetch("SELECT * FROM planes ORDER BY precio_mensual")
        return [dict(p) for p in planes]
    finally:
        await release_db(db)

@app.patch("/superadmin/planes/{plan_id}")
async def superadmin_actualizar_plan(request: Request, plan_id: str, datos: PlanUpdate):
    verificar_superadmin(request)
    db = await get_db()
    try:
        plan = await db.fetchrow("SELECT id FROM planes WHERE id = $1", plan_id)
        if not plan:
            raise HTTPException(status_code=404, detail="Plan no encontrado")
        if datos.nombre is not None:
            await db.execute("UPDATE planes SET nombre = $1 WHERE id = $2", datos.nombre, plan_id)
        if datos.precio_mensual is not None:
            await db.execute("UPDATE planes SET precio_mensual = $1 WHERE id = $2", datos.precio_mensual, plan_id)
        if datos.descripcion is not None:
            await db.execute("UPDATE planes SET descripcion = $1 WHERE id = $2", datos.descripcion, plan_id)
        return {"mensaje": "Plan actualizado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.get("/superadmin/stats")
async def superadmin_stats(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        stats = await db.fetchrow("""
            SELECT
                (SELECT COUNT(*) FROM tenants) as total_tenants,
                (SELECT COUNT(*) FROM tenants WHERE activo = TRUE) as tenants_activos,
                (SELECT COUNT(*) FROM tenants WHERE suscripcion_activa = TRUE) as tenants_pagos,
                (SELECT COUNT(*) FROM tenants WHERE activo = TRUE AND suscripcion_activa = FALSE AND trial_hasta >= CURRENT_DATE) as tenants_trial,
                (SELECT COUNT(*) FROM tenants WHERE activo = TRUE AND suscripcion_activa = FALSE AND trial_hasta < CURRENT_DATE) as tenants_vencidos,
                (SELECT COUNT(*) FROM usuarios) as total_usuarios,
                (SELECT COUNT(*) FROM tickets WHERE estado = 'abierto') as tickets_abiertos
        """)
        return dict(stats)
    finally:
        await release_db(db)

# ===== SUPERADMIN — FACTURACIÓN =====

class PagoCreate(BaseModel):
    tenant_id: str
    monto: float
    moneda: str = "USD"
    metodo: str
    referencia: Optional[str] = None
    notas: Optional[str] = None
    fecha: date
    registrado_por: Optional[str] = None

@app.get("/superadmin/facturacion")
async def superadmin_facturacion(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        ingreso = await db.fetchrow("""
            SELECT COALESCE(SUM(p.precio_mensual), 0) as ingreso_mensual
            FROM tenants t
            JOIN planes p ON t.plan_id = p.id
            WHERE t.suscripcion_activa = TRUE AND t.activo = TRUE
        """)
        alertas = await db.fetch("""
            SELECT t.nombre, t.email_admin, t.slug, t.trial_hasta,
                   (t.trial_hasta - CURRENT_DATE) as dias_restantes
            FROM tenants t
            WHERE t.activo = TRUE
              AND t.suscripcion_activa = FALSE
              AND t.trial_hasta BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '7 days'
            ORDER BY t.trial_hasta
        """)
        pagando = await db.fetch("""
            SELECT t.nombre, t.email_admin, t.slug, p.nombre as plan_nombre, p.precio_mensual as plan_precio
            FROM tenants t
            JOIN planes p ON t.plan_id = p.id
            WHERE t.suscripcion_activa = TRUE AND t.activo = TRUE
            ORDER BY t.nombre
        """)
        historial = await db.fetch("""
            SELECT pg.*, t.nombre as tenant_nombre, t.slug as tenant_slug
            FROM pagos pg
            JOIN tenants t ON pg.tenant_id = t.id
            ORDER BY pg.fecha DESC, pg.created_at DESC
            LIMIT 100
        """)
        return {
            "ingreso_mensual": float(ingreso["ingreso_mensual"]),
            "alertas_trial": [dict(a) for a in alertas],
            "tenants_pagando": [dict(p) for p in pagando],
            "historial_pagos": [dict(h) for h in historial]
        }
    finally:
        await release_db(db)

@app.post("/superadmin/pagos")
async def superadmin_registrar_pago(request: Request, datos: PagoCreate):
    verificar_superadmin(request)
    db = await get_db()
    try:
        tenant = await db.fetchrow("SELECT id, nombre FROM tenants WHERE id = $1", datos.tenant_id)
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant no encontrado")
        pago = await db.fetchrow(
            """INSERT INTO pagos (tenant_id, monto, moneda, metodo, referencia, notas, fecha, registrado_por)
               VALUES ($1, $2, $3, $4, $5, $6, $7, $8) RETURNING id""",
            datos.tenant_id, datos.monto, datos.moneda, datos.metodo,
            datos.referencia, datos.notas, datos.fecha, datos.registrado_por
        )
        return {"mensaje": f"Pago registrado para {tenant['nombre']}", "id": str(pago["id"])}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.delete("/superadmin/pagos/{pago_id}")
async def superadmin_eliminar_pago(request: Request, pago_id: str):
    verificar_superadmin(request)
    db = await get_db()
    try:
        result = await db.execute("DELETE FROM pagos WHERE id = $1", pago_id)
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="Pago no encontrado")
        return {"mensaje": "Pago eliminado"}
    finally:
        await release_db(db)

@app.get("/superadmin/facturacion/exportar")
async def superadmin_exportar_facturacion(request: Request, token: Optional[str] = None):
    if token:
        password_correcta = os.getenv("ADMIN_PASSWORD", "sL2#di!KBw")
        if token != password_correcta:
            raise HTTPException(status_code=401, detail="No autorizado")
    else:
        verificar_superadmin(request)
    db = await get_db()
    try:
        pagos = await db.fetch("""
            SELECT pg.fecha, t.nombre as tenant, t.slug, pg.monto, pg.moneda,
                   pg.metodo, pg.referencia, pg.notas, pg.registrado_por, pg.created_at
            FROM pagos pg
            JOIN tenants t ON pg.tenant_id = t.id
            ORDER BY pg.fecha DESC
        """)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturación"
        encabezados = ["Fecha", "Tenant", "Slug", "Monto", "Moneda", "Método", "Referencia", "Notas", "Registrado por", "Fecha registro"]
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=enc)
            celda.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            celda.fill = openpyxl.styles.PatternFill("solid", fgColor="2C3E50")
        for fila, p in enumerate(pagos, 2):
            ws.cell(row=fila, column=1, value=p["fecha"].strftime("%d/%m/%Y"))
            ws.cell(row=fila, column=2, value=p["tenant"])
            ws.cell(row=fila, column=3, value=p["slug"])
            ws.cell(row=fila, column=4, value=float(p["monto"]))
            ws.cell(row=fila, column=5, value=p["moneda"])
            ws.cell(row=fila, column=6, value=p["metodo"])
            ws.cell(row=fila, column=7, value=p["referencia"] or "")
            ws.cell(row=fila, column=8, value=p["notas"] or "")
            ws.cell(row=fila, column=9, value=p["registrado_por"] or "")
            ws.cell(row=fila, column=10, value=p["created_at"].strftime("%d/%m/%Y %H:%M"))
        anchos = [12, 25, 15, 10, 8, 15, 20, 25, 20, 18]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=facturacion.xlsx"}
        )
    finally:
        await release_db(db)

# ===== TICKETS DE SOPORTE =====

class TicketCreate(BaseModel):
    nombre: str
    email: str
    asunto: str
    descripcion: str
    prioridad: str = "media"
    slug: Optional[str] = None  # para identificar el tenant
    website: Optional[str] = None  # honeypot — debe venir vacío

class TicketRespuesta(BaseModel):
    respuesta: str
    estado: str

@app.post("/soporte/ticket")
@limiter.limit("5/hour")
async def crear_ticket(request: Request, datos: TicketCreate):
    """Endpoint público — no requiere autenticación. Máx 5 tickets por IP por hora."""
    # Validaciones de longitud
    if len(datos.nombre) > 100:
        raise HTTPException(status_code=400, detail="El nombre es demasiado largo")
    if len(datos.email) > 200:
        raise HTTPException(status_code=400, detail="El email es demasiado largo")
    if len(datos.asunto) > 200:
        raise HTTPException(status_code=400, detail="El asunto es demasiado largo")
    if len(datos.descripcion) > 3000:
        raise HTTPException(status_code=400, detail="La descripción es demasiado larga (máx 3000 caracteres)")
    if datos.prioridad not in ["alta", "media", "baja"]:
        raise HTTPException(status_code=400, detail="Prioridad inválida")
    # Honeypot — si viene con el campo trampa lleno, es un bot
    if datos.website:
        return {"mensaje": "Ticket creado correctamente"}  # respuesta falsa al bot
    db = await get_db()
    try:
        # Buscar tenant por slug si se provee
        tenant_id = None
        tenant_nombre = "Sin tenant"
        if datos.slug:
            tenant = await db.fetchrow(
                "SELECT id, nombre FROM tenants WHERE slug = $1",
                datos.slug
            )
            if tenant:
                tenant_id = tenant["id"]
                tenant_nombre = tenant["nombre"]

        ticket = await db.fetchrow(
            """INSERT INTO tickets (tenant_id, nombre, email, asunto, descripcion, prioridad)
               VALUES ($1, $2, $3, $4, $5, $6) RETURNING id, created_at""",
            tenant_id, datos.nombre, datos.email,
            datos.asunto, datos.descripcion, datos.prioridad
        )

        # Notificar al superadmin
        prioridad_emoji = {"alta": "🔴", "media": "🟡", "baja": "🟢"}.get(datos.prioridad, "🟡")
        enviar_email(
            "hola@reservatuespacio.com",
            f"{prioridad_emoji} Nuevo ticket de soporte — {datos.asunto}",
            f"""
            <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
                <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
                    <h2 style="color:#71D997; margin:0">Nuevo ticket de soporte</h2>
                </div>
                <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
                    <table style="width:100%; border-collapse:collapse">
                        <tr><td style="padding:8px; font-weight:bold; width:140px">Tenant:</td><td style="padding:8px">{tenant_nombre}</td></tr>
                        <tr><td style="padding:8px; font-weight:bold">Nombre:</td><td style="padding:8px">{datos.nombre}</td></tr>
                        <tr><td style="padding:8px; font-weight:bold">Email:</td><td style="padding:8px">{datos.email}</td></tr>
                        <tr><td style="padding:8px; font-weight:bold">Prioridad:</td><td style="padding:8px">{prioridad_emoji} {datos.prioridad.capitalize()}</td></tr>
                        <tr><td style="padding:8px; font-weight:bold">Asunto:</td><td style="padding:8px">{datos.asunto}</td></tr>
                        <tr><td style="padding:8px; font-weight:bold; vertical-align:top">Descripción:</td><td style="padding:8px">{datos.descripcion}</td></tr>
                    </table>
                    <div style="margin-top:20px; text-align:center">
                        <a href="https://claudiaacreativity.github.io/reserva-aulas/superadmin.html"
                           style="background:#71D997; color:#2C3E50; padding:12px 28px; border-radius:50px; text-decoration:none; font-weight:bold">
                            Ver en el panel superadmin →
                        </a>
                    </div>
                </div>
            </div>
            """
        )

        # Confirmar al tenant
        enviar_email(
            datos.email,
            f"✅ Recibimos tu consulta — {datos.asunto}",
            f"""
            <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
                <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
                    <h2 style="color:#71D997; margin:0">Recibimos tu consulta</h2>
                </div>
                <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
                    <p>Hola <b>{datos.nombre}</b>,</p>
                    <p>Recibimos tu consulta con el asunto <b>"{datos.asunto}"</b> y la vamos a revisar a la brevedad.</p>
                    <div style="background:white; border:1px solid #e0e0e0; border-radius:8px; padding:16px; margin:20px 0">
                        <p style="margin:0; color:#888; font-size:13px">Estado actual</p>
                        <p style="margin:4px 0 0; font-size:18px; font-weight:bold; color:#e67e22">🟡 Abierto</p>
                    </div>
                    <p style="color:#888; font-size:13px">Te vamos a notificar por email cuando tengamos una respuesta.</p>
                </div>
            </div>
            """
        )

        return {"mensaje": "Ticket creado correctamente", "id": str(ticket["id"])}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


@app.get("/soporte/tickets")
async def listar_tickets_tenant(email: str):
    """Devuelve los tickets de un email específico."""
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Email inválido")
    db = await get_db()
    try:
        tickets = await db.fetch(
            """SELECT id, asunto, descripcion, prioridad, estado, respuesta, created_at, updated_at
               FROM tickets WHERE email = $1
               ORDER BY created_at DESC""",
            email
        )
        return [dict(t) for t in tickets]
    finally:
        await release_db(db)

@app.get("/superadmin/tickets")
async def superadmin_listar_tickets(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        tickets = await db.fetch("""
            SELECT tk.*, t.nombre as tenant_nombre, t.slug as tenant_slug
            FROM tickets tk
            LEFT JOIN tenants t ON tk.tenant_id = t.id
            ORDER BY
                CASE tk.prioridad WHEN 'alta' THEN 1 WHEN 'media' THEN 2 ELSE 3 END,
                tk.created_at DESC
        """)
        return [dict(t) for t in tickets]
    finally:
        await release_db(db)

@app.patch("/superadmin/tickets/{ticket_id}")
async def superadmin_responder_ticket(request: Request, ticket_id: str, datos: TicketRespuesta):
    verificar_superadmin(request)
    db = await get_db()
    try:
        ticket = await db.fetchrow(
            "SELECT * FROM tickets WHERE id = $1", ticket_id
        )
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket no encontrado")

        await db.execute(
            """UPDATE tickets SET respuesta = $1, estado = $2, updated_at = NOW()
               WHERE id = $3""",
            datos.respuesta, datos.estado, ticket_id
        )

        estado_emoji = {"abierto": "🟡", "en_progreso": "🔵", "resuelto": "🟢"}.get(datos.estado, "🟡")
        estado_texto = {"abierto": "Abierto", "en_progreso": "En progreso", "resuelto": "Resuelto"}.get(datos.estado, datos.estado)

        # Notificar al tenant
        enviar_email(
            ticket["email"],
            f"{estado_emoji} Actualización de tu consulta — {ticket['asunto']}",
            f"""
            <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
                <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
                    <h2 style="color:#71D997; margin:0">Actualización de tu consulta</h2>
                </div>
                <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
                    <p>Hola <b>{ticket['nombre']}</b>,</p>
                    <p>Tu consulta <b>"{ticket['asunto']}"</b> fue actualizada.</p>
                    <div style="background:white; border:1px solid #e0e0e0; border-radius:8px; padding:16px; margin:20px 0">
                        <p style="margin:0; color:#888; font-size:13px">Estado actual</p>
                        <p style="margin:4px 0 0; font-size:18px; font-weight:bold; color:#2C3E50">{estado_emoji} {estado_texto}</p>
                    </div>
                    <div style="background:white; border-left:4px solid #71D997; padding:16px; border-radius:0 8px 8px 0; margin:20px 0">
                        <p style="margin:0; color:#888; font-size:13px">Respuesta</p>
                        <p style="margin:8px 0 0; color:#2C3E50">{datos.respuesta}</p>
                    </div>
                    <p style="color:#888; font-size:13px">Si tenés más preguntas podés escribirnos desde el formulario de soporte.</p>
                </div>
            </div>
            """
        )

        return {"mensaje": "Ticket actualizado y notificación enviada"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.get("/superadmin/tenant-por-email")
async def tenant_por_email(email: str):
    """Busca un tenant por el email del admin. Usado en el flujo de suscripción."""
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Email inválido")
    db = await get_db()
    try:
        tenant = await db.fetchrow(
            "SELECT id, nombre, slug, plan_id, suscripcion_activa, trial_hasta FROM tenants WHERE email_admin = $1 AND activo = TRUE",
            email
        )
        if not tenant:
            raise HTTPException(status_code=404, detail="No se encontró una cuenta con ese email")
        return dict(tenant)
    finally:
        await release_db(db)



# ===== MERCADOPAGO OAUTH (cobro por reserva) =====

MP_CLIENT_ID = "2324809191253560"  # Client ID de tu app ReservaTuEspacio

@app.get("/mp/oauth/autorizar")
async def mp_oauth_autorizar(request: Request):
    """Genera la URL para que el tenant autorice su cuenta de MP."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        redirect_uri = "https://reserva-aulas.onrender.com/mp/oauth/callback"
        url = (
            f"https://auth.mercadopago.com.ar/authorization"
            f"?client_id={MP_CLIENT_ID}"
            f"&response_type=code"
            f"&platform_id=mp"
            f"&redirect_uri={redirect_uri}"
            f"&state={tenant['id']}"
        )
        return {"url": url}
    finally:
        await release_db(db)

@app.get("/mp/oauth/callback")
async def mp_oauth_callback(code: str = None, state: str = None, error: str = None):
    """Recibe el código de autorización de MP y obtiene el access_token del tenant."""
    if error or not code or not state:
        return {"error": "Autorización cancelada o fallida"}

    client_secret = os.getenv("MP_CLIENT_SECRET", "")
    redirect_uri = "https://reserva-aulas.onrender.com/mp/oauth/callback"

    try:
        async with httpx.AsyncClient() as client:
            res = await client.post(
                "https://api.mercadopago.com/oauth/token",
                json={
                    "client_id": MP_CLIENT_ID,
                    "client_secret": client_secret,
                    "grant_type": "authorization_code",
                    "code": code,
                    "redirect_uri": redirect_uri
                },
                headers={"Content-Type": "application/json"}
            )

        if res.status_code != 200:
            return {"error": f"Error al obtener token: {res.text}"}

        data = res.json()
        access_token = data.get("access_token")
        refresh_token = data.get("refresh_token")
        mp_user_id = str(data.get("user_id", ""))

        db = await get_db()
        try:
            await db.execute(
                """UPDATE tenants SET mp_access_token = $1, mp_refresh_token = $2, mp_user_id = $3
                   WHERE id = $4""",
                access_token, refresh_token, mp_user_id, state
            )
        finally:
            await release_db(db)

        # Redirigir al panel admin con mensaje de éxito
        from fastapi.responses import RedirectResponse
        return RedirectResponse(
            url="https://reservatuespacio.com/admin.html?mp_conectado=1"
        )

    except Exception as e:
        return {"error": str(e)}

@app.delete("/mp/oauth/desconectar")
async def mp_oauth_desconectar(request: Request):
    """Desconecta la cuenta de MP del tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        await db.execute(
            """UPDATE tenants SET mp_access_token = NULL, mp_refresh_token = NULL, mp_user_id = NULL
               WHERE id = $1""",
            tenant["id"]
        )
        return {"mensaje": "Cuenta de MercadoPago desconectada"}
    finally:
        await release_db(db)

@app.get("/mp/estado")
async def mp_estado(request: Request):
    """Devuelve si el tenant tiene MP conectado y su configuración de cobro."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        return {
            "conectado": bool(tenant.get("mp_access_token")),
            "cobro_por_reserva": bool(tenant.get("mp_cobro_por_reserva")),
            "precio_reserva": float(tenant.get("mp_precio_reserva") or 0)
        }
    finally:
        await release_db(db)

class MPCobroConfig(BaseModel):
    cobro_por_reserva: bool
    precio_reserva: Optional[float] = None

@app.patch("/mp/cobro-config")
async def mp_cobro_config(request: Request, datos: MPCobroConfig):
    """Configura el cobro por reserva del tenant."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        if datos.cobro_por_reserva and not tenant.get("mp_access_token"):
            raise HTTPException(
                status_code=400,
                detail="Primero debés conectar tu cuenta de MercadoPago"
            )
        if datos.cobro_por_reserva and (not datos.precio_reserva or datos.precio_reserva <= 0):
            raise HTTPException(
                status_code=400,
                detail="Ingresá un precio por reserva válido"
            )
        await db.execute(
            """UPDATE tenants SET mp_cobro_por_reserva = $1, mp_precio_reserva = $2
               WHERE id = $3""",
            datos.cobro_por_reserva, datos.precio_reserva, tenant["id"]
        )
        return {"mensaje": "Configuración guardada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.post("/pagos/reserva/crear-preferencia")
async def crear_preferencia_reserva(request: Request):
    """Crea reserva con pending_payment y genera preferencia de pago MP."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)

        if not tenant.get("mp_cobro_por_reserva"):
            raise HTTPException(status_code=400, detail="Este tenant no tiene cobro por reserva activado")
        if not tenant.get("mp_access_token"):
            raise HTTPException(status_code=400, detail="El tenant no tiene MercadoPago conectado")

        body = await request.json()
        espacio_id        = body.get("espacio_id")
        espacio_nombre    = body.get("espacio_nombre", "Espacio")
        fecha             = body.get("fecha", "")
        hora_inicio       = body.get("hora_inicio", "")
        hora_fin          = body.get("hora_fin", "")
        usuario_id        = body.get("usuario_id")        # puede ser None (invitado)
        invitado_nombre   = body.get("invitado_nombre")
        invitado_email    = body.get("invitado_email")
        invitado_whatsapp = body.get("invitado_whatsapp")
        email_usuario     = invitado_email or body.get("email_usuario", "")
        nombre_usuario    = invitado_nombre or body.get("nombre_usuario", "")

        precio = float(tenant["mp_precio_reserva"])
        tc = await obtener_tipo_cambio(db)
        precio_ars = round(precio * tc, 2)
        expires_at = datetime.utcnow() + timedelta(minutes=15)

        # 1. Validar superposición antes de crear la reserva
        solapamiento = await db.fetchrow(
            """SELECT id FROM reservas
               WHERE aula_id = $1 AND fecha = $2
                 AND estado IN ('activa', 'pending_payment')
                 AND tenant_id = $3
                 AND hora_inicio < $4 AND hora_fin > $5""",
            espacio_id, fecha, tenant["id"],
            hora_fin, hora_inicio
        )
        if solapamiento:
            raise HTTPException(
                status_code=400,
                detail="Ese horario ya está reservado o está siendo procesado"
            )

        # 2. Crear reserva con estado pending_payment
        reserva_id = await db.fetchval(
            """INSERT INTO reservas
               (tenant_id, aula_id, usuario_id, invitado_nombre, invitado_email,
                invitado_whatsapp, fecha, hora_inicio, hora_fin,
                estado, expires_at, monto)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,'pending_payment',$10,$11)
               RETURNING id""",
            tenant["id"],
            espacio_id,
            usuario_id,
            invitado_nombre,
            invitado_email,
            invitado_whatsapp,
            fecha,
            hora_inicio,
            hora_fin,
            expires_at,
            precio
        )

        # 2. Crear preferencia en Mercado Pago
        preferencia_body = {
            "items": [{
                "title": f"Reserva {espacio_nombre} — {fecha} {hora_inicio}-{hora_fin}",
                "quantity": 1,
                "currency_id": "ARS",
                "unit_price": precio_ars
            }],
            "payer": {"email": email_usuario, "name": nombre_usuario},
            "back_urls": {
                "success": f"https://reservatuespacio.com/{tenant['slug']}?pago=ok&reserva_id={reserva_id}",
                "failure": f"https://reservatuespacio.com/{tenant['slug']}?pago=error&reserva_id={reserva_id}",
                "pending": f"https://reservatuespacio.com/{tenant['slug']}?pago=pendiente&reserva_id={reserva_id}"
            },
            "auto_return": "approved",
            "notification_url": "https://reserva-aulas.onrender.com/pagos/webhook-reserva",
            "external_reference": str(reserva_id),
            "statement_descriptor": tenant["nombre"][:22],
            "payment_methods": {"installments": 1}
        }

        async with httpx.AsyncClient() as client:
            res = await client.post(
                "https://api.mercadopago.com/checkout/preferences",
                json=preferencia_body,
                headers={
                    "Authorization": f"Bearer {tenant['mp_access_token']}",
                    "Content-Type": "application/json"
                }
            )

        if res.status_code not in (200, 201):
            # Si falla MP, cancelar la reserva creada
            await db.execute("UPDATE reservas SET estado='cancelada' WHERE id=$1", reserva_id)
            raise HTTPException(status_code=500, detail=f"Error MP: {res.text}")

        mp_data = res.json()
        preference_id = mp_data["id"]

        # 3. Guardar preference_id en la reserva
        await db.execute(
            "UPDATE reservas SET mp_preference_id=$1 WHERE id=$2",
            preference_id, reserva_id
        )

        return {
            "init_point": mp_data["init_point"],
            "preference_id": preference_id,
            "reserva_id": str(reserva_id),
            "precio_ars": precio_ars,
            "precio_usd": precio,
            "expires_at": expires_at.isoformat()
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await release_db(db)

@app.post("/pagos/webhook-reserva")
async def webhook_reserva(request: Request):
    """Webhook para pagos de reservas. Confirma o cancela la reserva automáticamente."""
    try:
        body = await request.json()
    except Exception:
        body = {}

    tipo = body.get("type") or request.query_params.get("type", "")
    topic = request.query_params.get("topic", "")

    if tipo != "payment" and topic != "payment":
        return {"status": "ignorado"}

    payment_id = body.get("data", {}).get("id") or request.query_params.get("id")
    if not payment_id:
        return {"status": "sin_payment_id"}

    db = await get_db()
    try:
        # Buscar el tenant dueño de este pago usando external_reference
        # Primero obtenemos el detalle del pago — necesitamos recorrer tenants con mp_access_token
        # para encontrar cuál es el dueño. Usamos external_reference (= reserva_id) como clave.

        # Obtener datos del pago iterando tokens de tenants activos con MP conectado
        pago = None
        tenant_row = None
        tenants_mp = await db.fetch(
            "SELECT id, mp_access_token FROM tenants WHERE mp_access_token IS NOT NULL AND activo = TRUE"
        )
        for t in tenants_mp:
            try:
                async with httpx.AsyncClient() as client:
                    res = await client.get(
                        f"https://api.mercadopago.com/v1/payments/{payment_id}",
                        headers={"Authorization": f"Bearer {t['mp_access_token']}"}
                    )
                if res.status_code == 200:
                    pago = res.json()
                    tenant_row = t
                    break
            except Exception:
                continue

        if not pago:
            return {"status": "pago_no_encontrado"}

        external_reference = pago.get("external_reference", "")
        status = pago.get("status", "")

        if not external_reference:
            return {"status": "sin_external_reference"}

        # Verificar que la reserva existe y pertenece al tenant
        reserva = await db.fetchrow(
            "SELECT * FROM reservas WHERE id=$1 AND tenant_id=$2",
            external_reference, tenant_row["id"]
        )
        if not reserva:
            return {"status": "reserva_no_encontrada"}

        if status == "approved":
            # Confirmar reserva
            await db.execute(
                """UPDATE reservas
                   SET estado='activa', mp_payment_id=$1, expires_at=NULL
                   WHERE id=$2""",
                str(payment_id), external_reference
            )
            # Enviar email de confirmación
            email_destino = reserva["invitado_email"]
            nombre_destino = reserva["invitado_nombre"] or "Cliente"
            if not email_destino and reserva["usuario_id"]:
                usuario = await db.fetchrow(
                    "SELECT email, nombre FROM usuarios WHERE id=$1",
                    reserva["usuario_id"]
                )
                if usuario:
                    email_destino = usuario["email"]
                    nombre_destino = usuario["nombre"]
            if email_destino:
                tenant_info = await db.fetchrow("SELECT nombre FROM tenants WHERE id=$1", tenant_row["id"])
                enviar_email(
                    email_destino,
                    "✅ Reserva confirmada",
                    f"""
                    <div style="font-family:Arial,sans-serif;max-width:500px;margin:auto">
                        <h2 style="color:#2C3E50">¡Reserva confirmada!</h2>
                        <p>Hola <b>{nombre_destino}</b>, tu pago fue aprobado y tu reserva está confirmada.</p>
                        <table style="border-collapse:collapse;width:100%;margin-top:16px">
                            <tr><td style="padding:10px 14px;font-weight:bold;color:#2C3E50">Fecha</td>
                                <td style="padding:10px 14px">{reserva['fecha']}</td></tr>
                            <tr><td style="padding:10px 14px;font-weight:bold;color:#2C3E50">Horario</td>
                                <td style="padding:10px 14px">{reserva['hora_inicio']} – {reserva['hora_fin']}</td></tr>
                            <tr><td style="padding:10px 14px;font-weight:bold;color:#2C3E50">Organización</td>
                                <td style="padding:10px 14px">{tenant_info['nombre']}</td></tr>
                        </table>
                        <p style="margin-top:20px;font-size:13px;color:#888;text-align:center">
                            <a href="https://reservatuespacio.com/faq-usuarios.html" style="color:#2C3E50;font-weight:bold">Preguntas frecuentes →</a>
                            &nbsp;&nbsp;|&nbsp;&nbsp;
                            <a href="https://reservatuespacio.com/soporte.html" style="color:#888">Centro de soporte →</a>
                        </p>
                    </div>
                    """
                )

        elif status in ("rejected", "cancelled"):
            await db.execute(
                "UPDATE reservas SET estado='cancelada' WHERE id=$1",
                external_reference
            )

        return {"status": "ok", "pago_status": status}

    except Exception as e:
        print(f"Error en webhook-reserva: {e}")
        return {"status": "error", "detalle": str(e)}
    finally:
        await release_db(db)

@app.get("/pagos/estado-reserva")
async def estado_reserva_pago(request: Request, preference_id: str):
    """Verifica si el pago de una reserva fue aprobado."""
    db = await get_db()
    try:
        tenant = await get_tenant(request, db)
        if not tenant.get("mp_access_token"):
            raise HTTPException(status_code=400, detail="MP no conectado")

        async with httpx.AsyncClient() as client:
            res = await client.get(
                f"https://api.mercadopago.com/checkout/preferences/{preference_id}",
                headers={"Authorization": f"Bearer {tenant['mp_access_token']}"}
            )

        if res.status_code != 200:
            return {"estado": "desconocido"}

        return {"estado": "ok", "preference": res.json()}
    finally:
        await release_db(db)

# ===== MERCADOPAGO (suscripciones GestionaTeIA) =====

class MPPreferenciaCreate(BaseModel):
    plan_id: str
    tenant_id: str
    email_admin: str
    nombre_tenant: str

@app.post("/pagos/crear-preferencia")
async def crear_preferencia_mp(request: Request, datos: MPPreferenciaCreate):
    """Crea una preferencia de pago en MercadoPago. Precio en USD convertido a ARS via BCRA."""
    db = await get_db()
    try:
        # Leer precio USD desde la tabla planes
        plan_db = await db.fetchrow("SELECT id, nombre, precio_mensual FROM planes WHERE id = $1", datos.plan_id.lower())
        if not plan_db:
            raise HTTPException(status_code=400, detail="Plan no válido")

        # Verificar que el tenant existe
        tenant = await db.fetchrow("SELECT id, nombre FROM tenants WHERE id = $1", datos.tenant_id)
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant no encontrado")

        access_token = os.getenv("MP_ACCESS_TOKEN")
        if not access_token:
            raise HTTPException(status_code=500, detail="MercadoPago no configurado")

        # Obtener tipo de cambio oficial y convertir a ARS
        tipo_cambio = await obtener_tipo_cambio(db)
        monto_ars = round(float(plan_db["precio_mensual"]) * tipo_cambio, 2)
        plan_nombre = plan_db["nombre"]

        preferencia_body = {
            "items": [
                {
                    "title": f"ReservaTuEspacio — Plan {plan_nombre}",
                    "description": f"Suscripción mensual plan {plan_nombre} para {datos.nombre_tenant} (USD {plan_db['precio_mensual']} al TC oficial)",
                    "quantity": 1,
                    "currency_id": "ARS",
                    "unit_price": monto_ars
                }
            ],
            "payer": {
                "email": datos.email_admin
            },
            "back_urls": {
                "success": f"https://reservatuespacio.com/pago-exitoso.html?tenant={datos.tenant_id}&plan={datos.plan_id}",
                "failure": f"https://reservatuespacio.com/landing.html?pago=fallido",
                "pending": f"https://reservatuespacio.com/pago-pendiente.html?tenant={datos.tenant_id}"
            },
            "auto_return": "approved",
            "notification_url": "https://reserva-aulas.onrender.com/pagos/webhook",
            "external_reference": f"{datos.tenant_id}|{datos.plan_id}",
            "statement_descriptor": "RESERVATUESPACIO",
            "payment_methods": {
                "excluded_payment_types": [],
                "installments": 1
            }
        }

        async with httpx.AsyncClient() as client:
            res = await client.post(
                "https://api.mercadopago.com/checkout/preferences",
                json=preferencia_body,
                headers={
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json"
                }
            )

        if res.status_code not in (200, 201):
            raise HTTPException(status_code=500, detail=f"Error al crear preferencia MP: {res.text}")

        mp_data = res.json()
        preference_id = mp_data["id"]

        # Guardar la preferencia en la base de datos
        await db.execute(
            """INSERT INTO mp_preferencias (tenant_id, preference_id, plan_id, monto)
               VALUES ($1, $2, $3, $4)""",
            datos.tenant_id, preference_id, datos.plan_id, monto_ars
        )

        return {
            "preference_id": preference_id,
            "init_point": mp_data["init_point"],         # URL producción
            "sandbox_init_point": mp_data["sandbox_init_point"]  # URL prueba
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await release_db(db)


@app.post("/pagos/webhook")
async def mp_webhook(request: Request):
    """MercadoPago llama a este endpoint cuando hay una novedad de pago."""
    try:
        body = await request.json()
    except Exception:
        body = {}

    # MP envía distintos tipos de notificaciones
    tipo = body.get("type") or request.query_params.get("type", "")
    topic = request.query_params.get("topic", "")

    # Solo procesamos pagos aprobados
    if tipo != "payment" and topic != "payment":
        return {"status": "ignorado"}

    payment_id = body.get("data", {}).get("id") or request.query_params.get("id")
    if not payment_id:
        return {"status": "sin payment_id"}

    access_token = os.getenv("MP_ACCESS_TOKEN")
    if not access_token:
        return {"status": "error_config"}

    try:
        # Consultar el pago a la API de MP para verificarlo
        async with httpx.AsyncClient() as client:
            res = await client.get(
                f"https://api.mercadopago.com/v1/payments/{payment_id}",
                headers={"Authorization": f"Bearer {access_token}"}
            )

        if res.status_code != 200:
            return {"status": "error_consulta_mp"}

        pago = res.json()
        estado = pago.get("status")
        external_ref = pago.get("external_reference", "")
        monto = float(pago.get("transaction_amount", 0))

        if not external_ref or "|" not in external_ref:
            return {"status": "referencia_invalida"}

        tenant_id, plan_id = external_ref.split("|", 1)

        db = await get_db()
        try:
            # Actualizar el registro de la preferencia
            await db.execute(
                """UPDATE mp_preferencias
                   SET estado = $1, mp_payment_id = $2, updated_at = NOW()
                   WHERE tenant_id = $3 AND plan_id = $4
                   AND estado = 'pendiente'""",
                estado, str(payment_id), tenant_id, plan_id
            )

            if estado == "approved":
                # Activar suscripción, calcular fecha de vencimiento (30 días) y resetear flags de avisos
                await db.execute(
                    """UPDATE tenants
                       SET suscripcion_activa = TRUE,
                           plan_id = $1,
                           suscripcion_vence = CURRENT_DATE + INTERVAL '30 days',
                           aviso_7_enviado = FALSE,
                           aviso_3_enviado = FALSE,
                           aviso_1_enviado = FALSE
                       WHERE id = $2""",
                    plan_id, tenant_id
                )

                # Registrar el pago en la tabla de pagos existente
                plan = {"nombre": plan_id}  # nombre de fallback
                tenant = await db.fetchrow(
                    "SELECT nombre, email_admin FROM tenants WHERE id = $1", tenant_id
                )

                if tenant:
                    await db.execute(
                        """INSERT INTO pagos (tenant_id, monto, moneda, metodo, referencia, notas, fecha, registrado_por)
                           VALUES ($1, $2, 'ARS', 'mercadopago', $3, $4, CURRENT_DATE, 'sistema')""",
                        tenant_id, monto, str(payment_id),
                        f"Pago automático plan {plan.get('nombre', plan_id)} vía MercadoPago"
                    )

                    # Email de confirmación al tenant
                    enviar_email(
                        tenant["email_admin"],
                        "✅ Pago confirmado — ReservaTuEspacio",
                        f"""
                        <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
                            <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
                                <h2 style="color:#71D997; margin:0">¡Pago confirmado!</h2>
                            </div>
                            <div style="background:#F9F9FB; padding:24px; border-radius:0 0 12px 12px">
                                <p>Hola <b>{tenant['nombre']}</b>,</p>
                                <p>Tu suscripción al plan <b>{plan.get('nombre', plan_id)}</b> fue activada correctamente.</p>
                                <table style="border-collapse:collapse; width:100%; background:white; border-radius:8px; overflow:hidden; margin:16px 0">
                                    <tr style="background:#f0f4f8">
                                        <td style="padding:10px 14px; font-weight:bold; color:#2C3E50; width:140px">Plan</td>
                                        <td style="padding:10px 14px; color:#4A5568">{plan.get('nombre', plan_id)}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Monto</td>
                                        <td style="padding:10px 14px; color:#4A5568">ARS {monto:,.0f}</td>
                                    </tr>
                                    <tr style="background:#f0f4f8">
                                        <td style="padding:10px 14px; font-weight:bold; color:#2C3E50">Referencia</td>
                                        <td style="padding:10px 14px; color:#4A5568">{payment_id}</td>
                                    </tr>
                                </table>
                                <p style="color:#4A5568">Tu acceso sigue activo en <b>reservatuespacio.com</b>.</p>
                                <p style="margin-top:20px; font-size:13px; color:#888; text-align:center">
                                    <a href="https://reservatuespacio.com/soporte.html" style="color:#888">¿Necesitás ayuda? Centro de soporte →</a>
                                </p>
                            </div>
                        </div>
                        """
                    )

        finally:
            await release_db(db)

        return {"status": "ok", "pago": estado}

    except Exception as e:
        print(f"Error en webhook MP: {e}")
        return {"status": "error", "detalle": str(e)}


@app.get("/pagos/estado")
async def estado_pago(tenant_id: str, plan_id: str):
    """Consulta el estado del último pago de un tenant para un plan."""
    db = await get_db()
    try:
        pref = await db.fetchrow(
            """SELECT estado, mp_payment_id, monto, created_at
               FROM mp_preferencias
               WHERE tenant_id = $1 AND plan_id = $2
               ORDER BY created_at DESC LIMIT 1""",
            tenant_id, plan_id
        )
        if not pref:
            return {"estado": "sin_pago"}
        return dict(pref)
    finally:
        await release_db(db)


@app.get("/pagos/planes")
async def listar_planes_precios():
    """Devuelve los planes con precio en USD y equivalente ARS al tipo de cambio oficial."""
    db = await get_db()
    try:
        planes = await db.fetch("SELECT id, nombre, precio_mensual FROM planes ORDER BY precio_mensual")
        tipo_cambio = await obtener_tipo_cambio(db)
        return [
            {
                "id": str(p["id"]),
                "nombre": p["nombre"],
                "precio_usd": float(p["precio_mensual"]),
                "precio_ars": round(float(p["precio_mensual"]) * tipo_cambio, 2),
                "tipo_cambio": tipo_cambio
            }
            for p in planes
        ]
    finally:
        await release_db(db)


# ===== COMUNICACIONES =====

class ComunicacionCreate(BaseModel):
    destinatarios: str  # "todos", "trial", "especifico"
    tenant_id: Optional[str] = None
    asunto: str
    mensaje: str

# ===== CONFIGURACIÓN GLOBAL =====

class ConfiguracionUpdate(BaseModel):
    valor: str

@app.get("/superadmin/configuracion")
async def superadmin_get_configuracion(request: Request):
    """Devuelve toda la configuración global del sistema."""
    verificar_superadmin(request)
    db = await get_db()
    try:
        rows = await db.fetch("SELECT clave, valor, descripcion, updated_at FROM configuracion_global ORDER BY clave")
        return [dict(r) for r in rows]
    finally:
        await release_db(db)

@app.patch("/superadmin/configuracion/{clave}")
async def superadmin_update_configuracion(request: Request, clave: str, datos: ConfiguracionUpdate):
    """Actualiza un valor de configuración global."""
    verificar_superadmin(request)
    db = await get_db()
    try:
        row = await db.fetchrow("SELECT clave FROM configuracion_global WHERE clave = $1", clave)
        if not row:
            raise HTTPException(status_code=404, detail="Clave de configuración no encontrada")
        await db.execute(
            "UPDATE configuracion_global SET valor = $1, updated_at = NOW() WHERE clave = $2",
            datos.valor, clave
        )
        return {"mensaje": f"Configuración '{clave}' actualizada correctamente", "valor": datos.valor}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)


# ===== COMUNICACIONES =====

@app.post("/superadmin/comunicaciones")
async def superadmin_enviar_comunicacion(request: Request, datos: ComunicacionCreate):
    verificar_superadmin(request)
    db = await get_db()
    try:
        if len(datos.asunto) > 200:
            raise HTTPException(status_code=400, detail="Asunto demasiado largo")
        if len(datos.mensaje) > 50000:
            raise HTTPException(status_code=400, detail="Mensaje demasiado largo")

        # Obtener destinatarios
        if datos.destinatarios == "especifico":
            if not datos.tenant_id:
                raise HTTPException(status_code=400, detail="Se requiere tenant_id")
            tenants = await db.fetch(
                "SELECT nombre, email_admin FROM tenants WHERE id = $1 AND activo = TRUE",
                datos.tenant_id
            )
        elif datos.destinatarios == "trial":
            tenants = await db.fetch(
                """SELECT nombre, email_admin FROM tenants
                   WHERE activo = TRUE AND suscripcion_activa = FALSE
                   AND trial_hasta >= CURRENT_DATE"""
            )
        else:  # todos
            tenants = await db.fetch(
                "SELECT nombre, email_admin FROM tenants WHERE activo = TRUE"
            )

        if not tenants:
            raise HTTPException(status_code=404, detail="No se encontraron destinatarios")

        total_enviados = 0
        for tenant in tenants:
            cuerpo = f"""
            <div style="font-family:Arial,sans-serif; max-width:600px; margin:0 auto">
                <div style="background:#2C3E50; padding:24px; border-radius:12px 12px 0 0">
                    <h2 style="color:#71D997; margin:0">reservatuespacio.com</h2>
                </div>
                <div style="background:#F9F9FB; padding:28px; border-radius:0 0 12px 12px">
                    <p style="color:#2C3E50; margin-bottom:20px">Hola <b>{tenant['nombre']}</b>,</p>
                    <div style="color:#4A5568; line-height:1.7">
                        {datos.mensaje}
                    </div>
                    <div style="border-top:1px solid #e0e0e0; margin-top:28px; padding-top:16px; text-align:center">
                        <a href="https://claudiaacreativity.github.io/reserva-aulas/soporte.html"
                           style="color:#888; font-size:12px; text-decoration:none">
                            ¿Necesitás ayuda? Centro de soporte →
                        </a>
                    </div>
                </div>
            </div>
            """
            enviar_email(tenant["email_admin"], datos.asunto, cuerpo)
            total_enviados += 1

        # Guardar en historial
        await db.execute(
            """INSERT INTO comunicaciones (destinatarios, asunto, mensaje, total_enviados)
               VALUES ($1, $2, $3, $4)""",
            datos.destinatarios, datos.asunto, datos.mensaje, total_enviados
        )

        return {"mensaje": f"Email enviado correctamente", "total_enviados": total_enviados}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await release_db(db)

@app.get("/superadmin/comunicaciones")
async def superadmin_historial_comunicaciones(request: Request):
    verificar_superadmin(request)
    db = await get_db()
    try:
        historial = await db.fetch(
            """SELECT * FROM comunicaciones ORDER BY created_at DESC LIMIT 50"""
        )
        return [dict(h) for h in historial]
    finally:
        await release_db(db)


# ===== RENOVACIÓN AUTOMÁTICA — PROCESAMIENTO DE VENCIMIENTOS =====

@app.post("/superadmin/procesar-vencimientos")
async def procesar_vencimientos(request: Request, token: Optional[str] = None):
    """
    Procesa vencimientos de suscripciones. Llamar diariamente vía cron job.
    Envía avisos 7, 3 y 1 día antes. Bloquea la cuenta el día del vencimiento.
    """
    password_correcta = os.getenv("ADMIN_PASSWORD", "sL2#di!KBw")
    # Acepta token por query param o por header X-Superadmin-Token
    token_recibido = token or request.headers.get("X-Superadmin-Token", "")
    if token_recibido != password_correcta:
        raise HTTPException(status_code=401, detail="No autorizado")

    db = await get_db()
    try:
        hoy = date.today()
        resumen = {
            "fecha": str(hoy),
            "avisos_7": [],
            "avisos_3": [],
            "avisos_1": [],
            "bloqueados": [],
            "errores": []
        }

        # Traer tenants con suscripción activa o vencida que tienen fecha de vencimiento
        tenants = await db.fetch("""
            SELECT id, nombre, email_admin, suscripcion_activa,
                   suscripcion_vence, aviso_7_enviado, aviso_3_enviado, aviso_1_enviado
            FROM tenants
            WHERE activo = TRUE
              AND suscripcion_vence IS NOT NULL
        """)

        for t in tenants:
            tid = t["id"]
            nombre = t["nombre"]
            email = t["email_admin"]
            vence = t["suscripcion_vence"]
            dias_restantes = (vence - hoy).days

            try:
                # ── BLOQUEO: ya venció ──────────────────────────────────────
                if dias_restantes < 0 and t["suscripcion_activa"]:
                    await db.execute(
                        "UPDATE tenants SET suscripcion_activa = FALSE WHERE id = $1",
                        tid
                    )
                    resumen["bloqueados"].append(nombre)
                    await enviar_email_cuenta_bloqueada(email, nombre, vence)

                # ── AVISO 1 día ─────────────────────────────────────────────
                elif dias_restantes == 1 and not t["aviso_1_enviado"]:
                    await db.execute(
                        "UPDATE tenants SET aviso_1_enviado = TRUE WHERE id = $1", tid
                    )
                    resumen["avisos_1"].append(nombre)
                    await enviar_email_aviso_vencimiento(email, nombre, 1, vence)

                # ── AVISO 3 días ────────────────────────────────────────────
                elif dias_restantes == 3 and not t["aviso_3_enviado"]:
                    await db.execute(
                        "UPDATE tenants SET aviso_3_enviado = TRUE WHERE id = $1", tid
                    )
                    resumen["avisos_3"].append(nombre)
                    await enviar_email_aviso_vencimiento(email, nombre, 3, vence)

                # ── AVISO 7 días ────────────────────────────────────────────
                elif dias_restantes == 7 and not t["aviso_7_enviado"]:
                    await db.execute(
                        "UPDATE tenants SET aviso_7_enviado = TRUE WHERE id = $1", tid
                    )
                    resumen["avisos_7"].append(nombre)
                    await enviar_email_aviso_vencimiento(email, nombre, 7, vence)

            except Exception as e:
                resumen["errores"].append(f"{nombre}: {str(e)}")

        # ── EXPIRACIÓN DE RESERVAS PENDING_PAYMENT ─────────────────────
        ahora = datetime.utcnow()
        reservas_expiradas = await db.fetch(
            """SELECT id FROM reservas
               WHERE estado = 'pending_payment'
                 AND expires_at IS NOT NULL
                 AND expires_at < $1""",
            ahora
        )
        ids_expirados = [str(r["id"]) for r in reservas_expiradas]
        if ids_expirados:
            await db.execute(
                """UPDATE reservas SET estado='expirada'
                   WHERE id = ANY($1::uuid[])""",
                ids_expirados
            )
        resumen["reservas_expiradas"] = len(ids_expirados)

        return {"ok": True, "resumen": resumen}

    finally:
        await release_db(db)
